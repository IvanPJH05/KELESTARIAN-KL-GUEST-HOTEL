from __future__ import annotations

import io
import json
import os
import re
import zipfile
from calendar import month_name, monthrange
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from flask import Flask, Response, jsonify, request
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.config["MAX_FORM_MEMORY_SIZE"] = 25 * 1024 * 1024

KELESTARIAN_START = date(2026, 1, 1)
KELESTARIAN_END = date(2026, 12, 31)


@dataclass
class Stay:
    bill_no: str | None
    registration_no: str | None
    folio_no: str | None
    guest_name: str | None
    room_no: str | None
    departure_date: date | None
    check_in_date: date | None
    check_out_date: date | None
    number_of_pax: int | None
    number_of_nights: int | None
    rate_type: str | None
    tariff_before_tax: Decimal
    tax_amount: Decimal
    food_amount: Decimal
    bar_amount: Decimal
    laundry_amount: Decimal
    call_amount: Decimal
    misc_amount: Decimal
    total_amount: Decimal
    amount_paid: Decimal
    payment_method: str | None
    flags: list[str]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def dec(value) -> Decimal:
    try:
        return Decimal(clean(value).replace(",", "") or "0")
    except Exception:
        return Decimal("0")


def num(value, default=0) -> int:
    try:
        return int(dec(value))
    except Exception:
        return default


def date_value(value, datemode=None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and datemode is not None:
        import xlrd

        try:
            return xlrd.xldate_as_datetime(value, datemode).date()
        except Exception:
            return None
    text = clean(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def money(value) -> str:
    amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{amount:,.2f}"


def supabase_configured() -> bool:
    return bool(os.environ.get("SUPABASE_URL") and os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))


def supabase_request(method: str, table: str, payload=None, query: dict | None = None, prefer: str | None = None):
    if not supabase_configured():
        raise RuntimeError("Supabase is not configured.")
    base = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    url = f"{base}/rest/v1/{table}"
    if query:
        url += "?" + urlencode(query)
    body = json.dumps(payload).encode("utf-8") if payload is not None else None
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    req = Request(url, data=body, headers=headers, method=method)
    try:
        with urlopen(req, timeout=20) as response:
            text = response.read().decode("utf-8")
            return json.loads(text) if text else None
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", "ignore")
        raise RuntimeError(f"Supabase request failed: {exc.code} {detail}") from exc


def supabase_select_all(table: str, query: dict, page_size: int = 1000) -> list[dict]:
    rows = []
    offset = 0
    while True:
        page_query = {**query, "limit": str(page_size), "offset": str(offset)}
        page = supabase_request("GET", table, query=page_query) or []
        rows.extend(page)
        if len(page) < page_size:
            return rows
        offset += page_size


def excel_rows(filename: str, data: bytes):
    lower = filename.lower()
    if lower.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_contents=data)
        for sheet in book.sheets():
            for row_idx in range(sheet.nrows):
                yield [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)], book.datemode
        return
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                yield list(row), None
        return
    raise ValueError("Please upload an .xls or .xlsx file.")


def parse_payment_method(text: str) -> str | None:
    match = re.search(r"\[\s*([^:\]]+)\s*:", clean(text))
    return normalize_payment_method(match.group(1).strip() if match else None)


def normalize_payment_method(value: str | None) -> str | None:
    text = clean(value).lower()
    if not text:
        return None
    if "online" in text or "booking" in text:
        return "Online Booking"
    if "cash" in text:
        return "Cash"
    if "atm" in text or "bank" in text or "transfer" in text:
        return "Bank / ATM"
    if "visa" in text or "master" in text or "american express" in text or "discover" in text or "card" in text:
        return "Card"
    if "check" in text or "cheque" in text:
        return "Cheque"
    if "security" in text or "deposit" in text:
        return "Security Deposit"
    if "reservation" in text:
        return "Reservation"
    if "extra" in text or "charge" in text:
        return "Extra Charge"
    return clean(value)


def parse_payment_amount(text: str) -> Decimal | None:
    match = re.search(r":\s*([0-9,]+(?:\.[0-9]+)?)\s*\]", clean(text))
    return dec(match.group(1)) if match else None


def parse_stays(filename: str, data: bytes) -> list[Stay]:
    stays = []
    rows = list(excel_rows(filename, data))
    index = 0
    while index < len(rows):
        row, datemode = rows[index]
        if len(row) < 17 or not clean(row[0]).upper().startswith("BN"):
            index += 1
            continue
        details = rows[index + 1][0] if index + 1 < len(rows) else []
        settlement = rows[index + 2][0] if index + 2 < len(rows) else []
        departure = date_value(row[3], datemode)
        nights = max(num(details[2] if len(details) > 2 else None, 0), 0)
        if not departure:
            index += 1
            continue
        check_in = departure - timedelta(days=nights) if nights else departure
        tariff = dec(row[5])
        tax = dec(row[10])
        food = dec(row[11])
        bar = dec(row[12])
        laundry = dec(row[13])
        call = dec(row[14])
        misc = dec(row[15])
        total = dec(row[16])
        paid = dec(settlement[0] if len(settlement) > 0 else None)
        payment_text = settlement[1] if len(settlement) > 1 else ""
        bracket_paid = parse_payment_amount(payment_text)
        if paid == Decimal("0") and bracket_paid is not None:
            paid = bracket_paid
        expected_total = tariff + tax + food + bar + laundry + call + misc
        flags = []
        if abs(paid - total) > Decimal("0.01"):
            flags.append("PAYMENT_MISMATCH")
        if abs(expected_total - total) > Decimal("0.01"):
            flags.append("TOTAL_MISMATCH")
        stays.append(
            Stay(
                bill_no=clean(row[0]) or None,
                registration_no=clean(row[1]) or None,
                folio_no=clean(row[2]) or None,
                guest_name=clean(row[4]).strip() or None,
                room_no=clean(details[0] if len(details) > 0 else None) or None,
                departure_date=departure,
                check_in_date=check_in,
                check_out_date=departure,
                number_of_pax=num(details[1] if len(details) > 1 else None, 0) or None,
                number_of_nights=nights or None,
                rate_type=clean(details[3] if len(details) > 3 else None) or None,
                tariff_before_tax=tariff,
                tax_amount=tax,
                food_amount=food,
                bar_amount=bar,
                laundry_amount=laundry,
                call_amount=call,
                misc_amount=misc,
                total_amount=total,
                amount_paid=paid,
                payment_method=parse_payment_method(payment_text),
                flags=flags,
            )
        )
        index += 4
    if not stays:
        raise ValueError("No guest rows were found. Check that the file is the Sales Bill Register export.")
    return sorted(stays, key=lambda s: (s.check_in_date, s.room_no or "", s.guest_name or ""))


def stay_record(stay: Stay) -> dict:
    return {
        "bill_no": stay.bill_no,
        "registration_no": stay.registration_no,
        "folio_no": stay.folio_no,
        "guest_name": stay.guest_name,
        "room_no": stay.room_no,
        "departure_date": stay.departure_date.isoformat(),
        "check_in_date": stay.check_in_date.isoformat(),
        "check_out_date": stay.check_out_date.isoformat(),
        "number_of_pax": stay.number_of_pax,
        "number_of_nights": stay.number_of_nights,
        "rate_type": stay.rate_type,
        "tariff_before_tax": float(stay.tariff_before_tax),
        "tax_amount": float(stay.tax_amount),
        "food_amount": float(stay.food_amount),
        "bar_amount": float(stay.bar_amount),
        "laundry_amount": float(stay.laundry_amount),
        "call_amount": float(stay.call_amount),
        "misc_amount": float(stay.misc_amount),
        "total_other_charges": float(stay.food_amount + stay.bar_amount + stay.laundry_amount + stay.call_amount + stay.misc_amount),
        "room_revenue": float(stay.tariff_before_tax),
        "total_amount": float(stay.total_amount),
        "amount_paid": float(stay.amount_paid),
        "payment_method": stay.payment_method,
        "flags": stay.flags,
    }


def flag_value(flags: list[str], key: str) -> str:
    prefix = f"{key}:"
    for flag in flags or []:
        text = clean(flag)
        if text.startswith(prefix):
            return text[len(prefix) :]
    return ""


def record_identity(record: dict) -> str:
    folio = verification_value(record.get("folio_no"))
    bill = verification_value(record.get("bill_no"))
    return folio or bill or f"{record.get('guest_name')}|{record.get('room_no')}|{record.get('check_in_date')}|{record.get('check_out_date')}"


def import_identity(record: dict) -> str:
    folio = verification_value(record.get("folio_no"))
    bill = verification_value(record.get("bill_no"))
    if folio:
        return f"FOLIO:{folio}"
    if bill:
        return f"BILL:{bill}"
    return "SIGNATURE:" + "|".join(
        verification_value(record.get(key))
        for key in ("guest_name", "room_no", "check_in_date", "check_out_date", "total_amount")
    )


def merge_stay_records(existing_records: list[dict], incoming_records: list[dict]) -> list[dict]:
    merged = {}
    for record in existing_records:
        merged[record_identity(record)] = record
    for record in incoming_records:
        merged[record_identity(record)] = record
    return list(merged.values())


def dedupe_import_records(records: list[dict]) -> tuple[list[dict], list[dict]]:
    deduped = {}
    duplicates = []
    for record in records:
        key = import_identity(record)
        previous = deduped.get(key)
        if previous:
            duplicates.append(
                {
                    "review_key": f"upload-duplicate|{key}",
                    "folio_no": record.get("folio_no"),
                    "incoming_bill_no": record.get("bill_no"),
                    "existing_folio_no": previous.get("folio_no"),
                    "existing_bill_no": previous.get("bill_no"),
                    "reason": "DUPLICATE_IN_UPLOAD",
                    "incoming_record": record,
                    "existing_record": previous,
                    "status": "pending",
                }
            )
        deduped[key] = record
    return list(deduped.values()), duplicates


def row_sort_key(record: dict) -> tuple[str, str]:
    return (clean(record.get("updated_at")), clean(record.get("id")))


def find_existing_matches(record: dict, existing_records: list[dict]) -> list[dict]:
    folio = verification_value(record.get("folio_no"))
    bill = verification_value(record.get("bill_no"))
    matches = []
    seen = set()
    for existing in existing_records:
        existing_id = existing.get("id")
        if existing_id in seen:
            continue
        if folio and verification_value(existing.get("folio_no")) == folio:
            matches.append(existing)
            seen.add(existing_id)
            continue
        if bill and verification_value(existing.get("bill_no")) == bill:
            matches.append(existing)
            seen.add(existing_id)
    return sorted(matches, key=row_sort_key, reverse=True)


def manual_import_match_key(record: dict) -> tuple[str, str]:
    return (clean(record.get("room_no")), clean(record.get("check_in_date")))


def manual_import_mismatch_reasons(manual: dict, incoming: dict) -> list[str]:
    reasons = []
    if clean(manual.get("room_no")) != clean(incoming.get("room_no")):
        reasons.append("ROOM")
    if clean(manual.get("check_in_date")) != clean(incoming.get("check_in_date")):
        reasons.append("CHECK_IN_DATE")
    manual_nights = int(dec(manual.get("number_of_nights")))
    incoming_nights = int(dec(incoming.get("number_of_nights")))
    if manual_nights != incoming_nights:
        reasons.append("NIGHTS")
    if abs(dec(manual.get("amount_paid")) - dec(incoming.get("amount_paid"))) > Decimal("0.01"):
        reasons.append("AMOUNT")
    return reasons


def is_manual_record(record: dict) -> bool:
    return "MANUAL_CHECK_IN" in (record.get("flags") or []) or verification_value(record.get("folio_no")).startswith("MANUAL-")


def merge_import_into_manual_record(manual: dict, incoming: dict, batch_id: str | None, now: str) -> dict:
    flags = list(manual.get("flags") or [])
    for flag in (
        "MATCHED_SALES_BILL_REGISTER",
        f"IMPORT_FOLIO:{verification_value(incoming.get('folio_no'))}",
        f"IMPORT_BILL:{verification_value(incoming.get('bill_no'))}",
    ):
        if flag and flag not in flags:
            flags.append(flag)
    return {
        **manual,
        "guest_name": incoming.get("guest_name") or manual.get("guest_name"),
        "registration_no": incoming.get("registration_no") or manual.get("registration_no"),
        "number_of_pax": incoming.get("number_of_pax") or manual.get("number_of_pax"),
        "number_of_nights": manual.get("number_of_nights") or incoming.get("number_of_nights"),
        "rate_type": incoming.get("rate_type") or manual.get("rate_type"),
        "flags": flags,
        "import_batch_id": batch_id,
        "updated_at": now,
    }


def queue_review_records(records: list[dict], batch_id: str | None, filename: str) -> None:
    if not records:
        return
    now = datetime.utcnow().isoformat() + "Z"
    for record in records:
        record["import_batch_id"] = batch_id
        record["source_filename"] = filename
        record["updated_at"] = now
    for index in range(0, len(records), 500):
        supabase_request(
            "POST",
            "guest_stay_review_queue",
            records[index : index + 500],
            query={"on_conflict": "review_key"},
            prefer="resolution=merge-duplicates",
        )


def verification_value(value) -> str:
    return clean(value).upper()


def unique_verification_values(values) -> list[str]:
    seen = set()
    result = []
    for value in values:
        normalized = verification_value(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def classify_verified_stays(stays: list[Stay], existing_records: list[dict]) -> tuple[list[dict], list[dict]]:
    existing_by_folio = {
        verification_value(row.get("folio_no")): row
        for row in existing_records
        if verification_value(row.get("folio_no"))
    }
    existing_by_bill = {
        verification_value(row.get("bill_no")): row
        for row in existing_records
        if verification_value(row.get("bill_no"))
    }
    accepted_by_folio = {}
    accepted_bill_to_folio = {}
    reviews = []
    for stay in stays:
        incoming = stay_record(stay)
        folio = verification_value(stay.folio_no)
        bill = verification_value(stay.bill_no)
        incoming["folio_no"] = folio or None
        incoming["bill_no"] = bill or None
        folio_match = accepted_by_folio.get(folio) or existing_by_folio.get(folio)
        bill_match = existing_by_bill.get(bill)
        accepted_bill_folio = accepted_bill_to_folio.get(bill)
        existing_bill_folio = verification_value(bill_match.get("folio_no")) if bill_match else ""
        reason = ""
        existing = None
        if not folio or not bill:
            reason = "MISSING_FOLIO_OR_BILL"
        elif folio_match and verification_value(folio_match.get("bill_no")) != bill:
            reason = "FOLIO_MATCHES_DIFFERENT_BILL"
            existing = folio_match
        elif accepted_bill_folio and accepted_bill_folio != folio:
            reason = "BILL_MATCHES_DIFFERENT_FOLIO"
            existing = accepted_by_folio.get(accepted_bill_folio)
        elif bill_match and existing_bill_folio != folio:
            reason = "BILL_MATCHES_DIFFERENT_FOLIO"
            existing = bill_match
        if reason:
            reviews.append(
                {
                    "review_key": f"{folio or '(missing)'}|{bill or '(missing)'}",
                    "folio_no": folio or None,
                    "incoming_bill_no": bill or None,
                    "existing_folio_no": verification_value((existing or {}).get("folio_no")) or None,
                    "existing_bill_no": verification_value((existing or {}).get("bill_no")) or None,
                    "reason": reason,
                    "incoming_record": incoming,
                    "existing_record": existing,
                    "status": "pending",
                }
            )
            continue
        accepted_by_folio[folio] = incoming
        accepted_bill_to_folio[bill] = folio
    return list(accepted_by_folio.values()), reviews


def save_verified_stay_rows(rows: list[dict], existing_records: list[dict]) -> None:
    existing_by_folio = {
        verification_value(row.get("folio_no")): row
        for row in existing_records
        if verification_value(row.get("folio_no"))
    }
    existing_by_bill = {
        verification_value(row.get("bill_no")): row
        for row in existing_records
        if verification_value(row.get("bill_no"))
    }
    updates = []
    inserts = []
    for row in rows:
        folio = verification_value(row.get("folio_no"))
        bill = verification_value(row.get("bill_no"))
        existing = existing_by_folio.get(folio) or existing_by_bill.get(bill)
        if not existing:
            inserts.append(row)
            continue
        if existing.get("id"):
            updates.append({**row, "id": existing["id"]})
        else:
            updates.append(row)
    for index in range(0, len(updates), 500):
        supabase_request(
            "POST",
            "guest_stays",
            updates[index : index + 500],
            query={"on_conflict": "id"},
            prefer="resolution=merge-duplicates,return=minimal",
        )
    for index in range(0, len(inserts), 500):
        supabase_request("POST", "guest_stays", inserts[index : index + 500], prefer="return=minimal")


def save_import_to_supabase(filename: str, stays: list[Stay], sales: list[dict], summary_data: dict) -> dict:
    if not supabase_configured():
        return {"saved": False, "accepted_count": 0, "review_count": 0}
    batch = supabase_request(
        "POST",
        "hotel_import_batches",
        {
            "source_filename": filename,
            "stay_count": len(stays),
            "sale_row_count": len(sales),
            "total_sales": str(dec(summary_data.get("total_sales"))),
            "total_kelestarian": str(dec(summary_data.get("total_kelestarian"))),
        },
        query={"select": "id"},
        prefer="return=representation",
    )
    batch_id = batch[0]["id"] if batch else None
    existing_records = load_matching_stays_from_supabase(stays)
    incoming_rows, reviews = dedupe_import_records([stay_record(stay) for stay in stays])
    inserted = 0
    updated = 0
    manual_matched = 0
    merged_duplicates = 0
    rows_to_insert = []
    row_ids_to_delete = []
    now = datetime.utcnow().isoformat() + "Z"
    existing_by_id = {row.get("id"): row for row in existing_records if row.get("id")}
    incoming_dates = [date_value(row.get("check_in_date")) for row in incoming_rows]
    incoming_dates = [value for value in incoming_dates if value]
    manual_by_room_date = {}
    if filename != "Manual daily check-in" and incoming_dates:
        manual_records = [
            row
            for row in load_stays_by_checkin_range_from_supabase(min(incoming_dates), max(incoming_dates))
            if is_manual_record(row)
        ]
        for row in manual_records:
            manual_by_room_date[manual_import_match_key(row)] = row
    for row in incoming_rows:
        row["folio_no"] = verification_value(row.get("folio_no")) or None
        row["bill_no"] = verification_value(row.get("bill_no")) or None
        row["import_batch_id"] = batch_id
        row["updated_at"] = now
        manual_match = manual_by_room_date.get(manual_import_match_key(row))
        if manual_match and manual_match.get("id"):
            mismatch_reasons = manual_import_mismatch_reasons(manual_match, row)
            if mismatch_reasons:
                reviews.append(
                    {
                        "review_key": f"manual-import-mismatch|{manual_match.get('id')}|{row.get('folio_no') or row.get('bill_no')}",
                        "folio_no": row.get("folio_no"),
                        "incoming_bill_no": row.get("bill_no"),
                        "existing_folio_no": manual_match.get("folio_no"),
                        "existing_bill_no": manual_match.get("bill_no"),
                        "reason": "MANUAL_IMPORT_MISMATCH:" + ",".join(mismatch_reasons),
                        "incoming_record": row,
                        "existing_record": manual_match,
                        "status": "pending",
                    }
                )
                continue
            manual_matched += 1
            updated += 1
            rows_to_insert.append(merge_import_into_manual_record(manual_match, row, batch_id, now))
            existing_by_id.pop(manual_match.get("id"), None)
            continue
        matches = find_existing_matches(row, list(existing_by_id.values()))
        if not matches:
            inserted += 1
            rows_to_insert.append(row)
            continue
        updated += 1
        rows_to_insert.append(row)
        for matched in matches[1:]:
            matched_id = matched.get("id")
            if not matched_id:
                continue
            reviews.append(
                {
                    "review_key": f"existing-duplicate-merged|{matched_id}",
                    "folio_no": row.get("folio_no"),
                    "incoming_bill_no": row.get("bill_no"),
                    "existing_folio_no": matched.get("folio_no"),
                    "existing_bill_no": matched.get("bill_no"),
                    "reason": "EXISTING_DUPLICATE_MERGED",
                    "incoming_record": row,
                    "existing_record": matched,
                    "status": "pending",
                }
            )
            merged_duplicates += 1
            row_ids_to_delete.append(matched_id)
            existing_by_id.pop(matched_id, None)
    for index in range(0, len(row_ids_to_delete), 100):
        chunk = row_ids_to_delete[index : index + 100]
        supabase_request(
            "DELETE",
            "guest_stays",
            query={"id": "in.(" + ",".join(chunk) + ")"},
            prefer="return=minimal",
        )
    for index in range(0, len(rows_to_insert), 500):
        chunk = rows_to_insert[index : index + 500]
        supabase_request(
            "POST",
            "guest_stays",
            chunk,
            query={"on_conflict": "folio_no"},
            prefer="resolution=merge-duplicates,return=minimal",
        )
    review_error = ""
    try:
        queue_review_records(reviews, batch_id, filename)
    except Exception as exc:
        review_error = str(exc)
    return {
        "saved": True,
        "accepted_count": len(incoming_rows),
        "inserted_count": inserted,
        "updated_count": updated,
        "manual_matched_count": manual_matched,
        "merged_duplicate_count": merged_duplicates,
        "review_count": len(reviews),
        "review_error": review_error,
    }


def load_stays_from_supabase() -> list[dict]:
    if not supabase_configured():
        return []
    return supabase_select_all(
        "guest_stays",
        query={
            "select": "*",
            "order": "check_in_date.asc,room_no.asc,guest_name.asc",
        },
    )


def load_stays_by_checkout_range_from_supabase(start: date, end: date) -> list[dict]:
    if not supabase_configured():
        return []
    return supabase_select_all(
        "guest_stays",
        query={
            "select": "*",
            "and": f"(check_out_date.gte.{start.isoformat()},check_out_date.lte.{end.isoformat()})",
            "order": "check_out_date.asc,room_no.asc,guest_name.asc",
        },
    )


def load_stays_by_checkin_range_from_supabase(start: date, end: date) -> list[dict]:
    if not supabase_configured():
        return []
    return supabase_select_all(
        "guest_stays",
        query={
            "select": "*",
            "and": f"(check_in_date.gte.{start.isoformat()},check_in_date.lte.{end.isoformat()})",
            "order": "check_in_date.asc,room_no.asc,guest_name.asc",
        },
    )


def load_matching_stays_from_supabase(stays: list[Stay]) -> list[dict]:
    if not supabase_configured():
        return []
    found = {}
    lookups = [
        ("folio_no", unique_verification_values(stay.folio_no for stay in stays)),
        ("bill_no", unique_verification_values(stay.bill_no for stay in stays)),
    ]
    for column, values in lookups:
        for index in range(0, len(values), 100):
            chunk = values[index : index + 100]
            if not chunk:
                continue
            rows = supabase_select_all(
                "guest_stays",
                query={
                    "select": "*",
                    column: "in.(" + ",".join(chunk) + ")",
                },
            )
            for row in rows:
                found[row.get("id") or record_identity(row)] = row
    return list(found.values())


def load_review_queue_from_supabase() -> list[dict]:
    if not supabase_configured():
        return []
    return supabase_request(
        "GET",
        "guest_stay_review_queue",
        query={
            "select": "id,folio_no,incoming_bill_no,existing_folio_no,existing_bill_no,reason,incoming_record,existing_record,source_filename,status,created_at,updated_at",
            "status": "eq.pending",
            "order": "updated_at.desc",
            "limit": "1000",
        },
    ) or []


def load_import_batches_from_supabase() -> list[dict]:
    if not supabase_configured():
        return []
    return supabase_select_all(
        "hotel_import_batches",
        query={
            "select": "id,source_filename,imported_at,stay_count,sale_row_count,total_sales,total_kelestarian",
            "order": "imported_at.desc",
        },
    )


def records_to_stays(records: list[dict]) -> list[Stay]:
    stays = []
    for item in records:
        check_in = date_value(item.get("check_in_date"))
        check_out = date_value(item.get("check_out_date"))
        if not check_in or not check_out:
            continue
        nights = max(num(item.get("number_of_nights"), 0), 0)
        total = dec(item.get("total_amount"))
        stays.append(
            Stay(
                bill_no=clean(item.get("bill_no")) or None,
                registration_no=clean(item.get("registration_no")) or None,
                folio_no=clean(item.get("folio_no")) or None,
                guest_name=clean(item.get("guest_name")) or None,
                room_no=clean(item.get("room_no")),
                departure_date=check_out,
                check_in_date=check_in,
                check_out_date=check_out,
                number_of_pax=num(item.get("number_of_pax"), 0) or None,
                number_of_nights=nights or None,
                rate_type=clean(item.get("rate_type")) or None,
                tariff_before_tax=dec(item.get("tariff_before_tax")),
                tax_amount=dec(item.get("tax_amount")),
                food_amount=dec(item.get("food_amount")),
                bar_amount=dec(item.get("bar_amount")),
                laundry_amount=dec(item.get("laundry_amount")),
                call_amount=dec(item.get("call_amount")),
                misc_amount=dec(item.get("misc_amount")),
                total_amount=total,
                amount_paid=dec(item.get("amount_paid")),
                payment_method=normalize_payment_method(item.get("payment_method")),
                flags=item.get("flags") or [],
            )
        )
    if not stays:
        raise ValueError("No imported stays found. Please import the Excel file again.")
    return sorted(stays, key=lambda s: (s.check_in_date, s.room_no or "", s.guest_name or ""))


def expanded_sales(stays: list[Stay], fee_rate: Decimal) -> list[dict]:
    rows = []
    for stay in stays:
        if not stay.check_in_date or not stay.number_of_nights:
            continue
        if stay.check_in_date < KELESTARIAN_START or stay.check_in_date > KELESTARIAN_END:
            continue
        stay_last_night = stay.check_in_date + timedelta(days=stay.number_of_nights - 1)
        charge_start = stay.check_in_date
        charge_end = min(stay_last_night, KELESTARIAN_END)
        if charge_start > charge_end:
            continue
        chargeable_nights = (charge_end - charge_start).days + 1
        total_kelestarian = fee_rate * Decimal(chargeable_nights)
        paid_date = stay.check_in_date.strftime("%d/%m/%Y")
        for offset in range(chargeable_nights):
            day = charge_start + timedelta(days=offset)
            original_offset = (day - stay.check_in_date).days
            is_charge_start_day = offset == 0
            rows.append(
                {
                    "date": day.isoformat(),
                    "display_date": day.strftime("%d/%m/%Y"),
                    "room_no": stay.room_no,
                    "guest_name": stay.guest_name,
                    "stay_progress": f"{original_offset + 1}/{stay.number_of_nights}",
                    "nights": chargeable_nights,
                    "multi_night": chargeable_nights > 1,
                    "price": money(stay.amount_paid if is_charge_start_day else Decimal("0")),
                    "total_paid": money(stay.amount_paid),
                    "amount_before_tax": money(stay.tariff_before_tax if is_charge_start_day else Decimal("0")),
                    "room_revenue": money(stay.tariff_before_tax if is_charge_start_day else Decimal("0")),
                    "kelestarian": money(total_kelestarian if is_charge_start_day else Decimal("0")),
                    "payment_status": "Collected" if is_charge_start_day else f"Paid on {paid_date}",
                    "paid_date": charge_start.isoformat(),
                    "is_paid_continuation": not is_charge_start_day,
                    "bill_no": stay.bill_no,
                    "payment_method": stay.payment_method,
                    "rate_type": stay.rate_type,
                    "flags": stay.flags,
                }
            )
    return sorted(rows, key=lambda r: (r["date"], r["room_no"], r["guest_name"]))


def ledger_sales(stays: list[Stay], fee_rate: Decimal) -> list[dict]:
    rows = []
    for stay in stays:
        if not stay.check_in_date or not stay.number_of_nights:
            continue
        stay_last_night = stay.check_in_date + timedelta(days=stay.number_of_nights - 1)
        chargeable_nights = 0
        if KELESTARIAN_START <= stay.check_in_date <= KELESTARIAN_END:
            charge_end = min(stay_last_night, KELESTARIAN_END)
            chargeable_nights = max((charge_end - stay.check_in_date).days + 1, 0)
        total_kelestarian = fee_rate * Decimal(chargeable_nights)
        deposit_amount = dec(flag_value(stay.flags, "DEPOSIT"))
        kelestarian_payment_method = flag_value(stay.flags, "KELESTARIAN_PAYMENT") or stay.payment_method or ""
        deposit_payment_method = flag_value(stay.flags, "DEPOSIT_PAYMENT")
        registration_label = flag_value(stay.flags, "REGISTRATION") or stay.registration_no or ""
        stay_type_label = flag_value(stay.flags, "STAY_TYPE") or stay.rate_type or ""
        input_time = flag_value(stay.flags, "INPUT_TIME")
        remark = flag_value(stay.flags, "REMARK")
        paid_date = stay.check_in_date.strftime("%d/%m/%Y")
        for offset in range(stay.number_of_nights):
            day = stay.check_in_date + timedelta(days=offset)
            is_check_in_day = offset == 0
            rows.append(
                {
                    "date": day.isoformat(),
                    "display_date": day.strftime("%d/%m/%Y"),
                    "input_time": input_time,
                    "folio_no": stay.folio_no,
                    "room_no": stay.room_no,
                    "guest_name": stay.guest_name,
                    "registration": registration_label,
                    "stay_type_label": stay_type_label,
                    "remark": remark,
                    "check_in_date": stay.check_in_date.isoformat(),
                    "check_out_date": stay.check_out_date.isoformat() if stay.check_out_date else "",
                    "stay_progress": f"{offset + 1}/{stay.number_of_nights}",
                    "nights": stay.number_of_nights,
                    "multi_night": stay.number_of_nights > 1,
                    "price": money(stay.amount_paid if is_check_in_day else Decimal("0")),
                    "total_paid": money(stay.amount_paid),
                    "amount_before_tax": money(stay.tariff_before_tax if is_check_in_day else Decimal("0")),
                    "room_revenue": money(stay.tariff_before_tax if is_check_in_day else Decimal("0")),
                    "kelestarian": money(total_kelestarian if is_check_in_day else Decimal("0")),
                    "payment_status": "Collected" if is_check_in_day else f"Paid on {paid_date}",
                    "paid_date": stay.check_in_date.isoformat(),
                    "is_paid_continuation": not is_check_in_day,
                    "deposit": money(deposit_amount if is_check_in_day else Decimal("0")) if deposit_amount else "",
                    "deposit_payment_method": deposit_payment_method if is_check_in_day else "",
                    "kelestarian_payment_method": kelestarian_payment_method if is_check_in_day and total_kelestarian else "",
                    "bill_no": stay.bill_no,
                    "payment_method": stay.payment_method,
                    "rate_type": stay.rate_type,
                    "flags": stay.flags,
                }
            )
    return sorted(rows, key=lambda r: (r["date"], r["room_no"], r["guest_name"]))


def summary(stays: list[Stay], fee_rate: Decimal) -> dict:
    sales = expanded_sales(stays, fee_rate)
    total_fee = sum((dec(row["kelestarian"]) for row in sales), Decimal("0"))
    total_sales = sum((s.total_amount for s in stays), Decimal("0"))
    days = sorted({row["date"] for row in sales})
    daily_revenue = defaultdict(lambda: {"rooms": 0, "total_amount": Decimal("0"), "kelestarian": Decimal("0")})
    payment_breakdown = defaultdict(lambda: {"count": 0, "total_amount": Decimal("0")})
    room_revenue_report = defaultdict(lambda: {"nights": 0, "room_revenue": Decimal("0"), "total_amount": Decimal("0")})
    occupancy_by_date = defaultdict(set)
    revenue_by_rate_type = defaultdict(lambda: {"count": 0, "total_amount": Decimal("0")})
    issues = []
    for row in sales:
        day = row["display_date"]
        daily_revenue[day]["rooms"] += 1
        daily_revenue[day]["total_amount"] += dec(row["price"])
        daily_revenue[day]["kelestarian"] += dec(row["kelestarian"])
        occupancy_by_date[day].add(row["room_no"])
    for stay in stays:
        payment = stay.payment_method or "Unknown"
        payment_breakdown[payment]["count"] += 1
        payment_breakdown[payment]["total_amount"] += stay.amount_paid
        room_key = stay.room_no or "Unknown"
        room_revenue_report[room_key]["nights"] += stay.number_of_nights or 0
        room_revenue_report[room_key]["room_revenue"] += stay.tariff_before_tax
        room_revenue_report[room_key]["total_amount"] += stay.total_amount
        rate_type = stay.rate_type or "Unknown"
        revenue_by_rate_type[rate_type]["count"] += 1
        revenue_by_rate_type[rate_type]["total_amount"] += stay.total_amount
        if stay.flags:
            issues.append(
                {
                    "bill_no": stay.bill_no,
                    "guest_name": stay.guest_name,
                    "room_no": stay.room_no,
                    "flags": stay.flags,
                    "total_amount": money(stay.total_amount),
                    "amount_paid": money(stay.amount_paid),
                }
            )
    total_nights = sum((s.number_of_nights or 0) for s in stays)
    average_length = Decimal(total_nights) / Decimal(len(stays)) if stays else Decimal("0")
    top_guests = sorted(stays, key=lambda s: s.total_amount, reverse=True)[:10]
    return {
        "stays": len(stays),
        "sale_rows": len(sales),
        "days": len(days),
        "first_date": days[0] if days else "",
        "last_date": days[-1] if days else "",
        "total_sales": money(total_sales),
        "total_kelestarian": money(total_fee),
        "daily_revenue_summary": [
            {"date": key, "rooms": value["rooms"], "total_amount": money(value["total_amount"]), "kelestarian": money(value["kelestarian"])}
            for key, value in daily_revenue.items()
        ],
        "payment_method_breakdown": [
            {"payment_method": key, "count": value["count"], "total_amount": money(value["total_amount"])}
            for key, value in sorted(payment_breakdown.items())
        ],
        "room_revenue_report": [
            {"room_no": key, "nights": value["nights"], "room_revenue": money(value["room_revenue"]), "total_amount": money(value["total_amount"])}
            for key, value in sorted(room_revenue_report.items())
        ],
        "occupancy_by_date": [
            {"date": key, "occupied_rooms": len(value)}
            for key, value in occupancy_by_date.items()
        ],
        "average_length_of_stay": money(average_length),
        "revenue_before_tax": money(sum((s.tariff_before_tax for s in stays), Decimal("0"))),
        "total_tax_collected": money(sum((s.tax_amount for s in stays), Decimal("0"))),
        "top_guests_by_spend": [
            {"guest_name": s.guest_name, "room_no": s.room_no, "total_amount": money(s.total_amount)}
            for s in top_guests
        ],
        "revenue_by_rate_type": [
            {"rate_type": key, "count": value["count"], "total_amount": money(value["total_amount"])}
            for key, value in sorted(revenue_by_rate_type.items())
        ],
        "data_quality_issues": issues,
    }


def kelestarian_stays(stays: list[Stay]) -> list[Stay]:
    return [stay for stay in stays if stay.check_in_date and KELESTARIAN_START <= stay.check_in_date <= KELESTARIAN_END]


def manual_payment_method(value: str) -> str:
    allowed = {
        "cash": "Cash",
        "qr": "QR",
        "card": "Card",
        "online": "Online",
        "transfer": "Transfer",
    }
    key = clean(value).lower()
    if key not in allowed:
        raise ValueError("Please choose a valid payment method.")
    return allowed[key]


def manual_stay_from_form(form) -> Stay:
    room = clean(form.get("room"))
    if not room:
        raise ValueError("Room is required.")
    registration = clean(form.get("registration")).lower() or "walk in"
    if registration not in {"walk in", "online", "extension"}:
        raise ValueError("Please choose a valid registration type.")
    check_in = date.fromisoformat(clean(form.get("check_in_date")) or date.today().isoformat())
    stay_type = clean(form.get("stay_type")) or "night"
    if stay_type == "3_hour":
        nights = 1
        rate_type = "3 Hour"
        duration_label = "3 hour"
    else:
        nights = max(num(form.get("stay_duration"), 1), 1)
        rate_type = registration.upper()
        duration_label = f"{nights} night"
    check_out = date.fromisoformat(clean(form.get("check_out_date")) or (check_in + timedelta(days=nights)).isoformat())
    amount = dec(form.get("amount_paid"))
    if amount <= 0:
        raise ValueError("Room payment must be more than RM0.")
    payment_method = manual_payment_method(form.get("payment_method"))
    kelestarian_payment = manual_payment_method(form.get("kelestarian_payment_method") or form.get("payment_method"))
    deposit_payment = manual_payment_method(form.get("deposit_payment_method") or form.get("payment_method"))
    deposit_amount = dec(form.get("deposit_amount") or "50")
    input_time = datetime.now(UTC)
    timestamp = input_time.strftime("%Y%m%d%H%M%S")
    identity = f"MANUAL-{check_in.strftime('%Y%m%d')}-{room}-{timestamp}"
    flags = [
        "MANUAL_CHECK_IN",
        f"INPUT_TIME:{input_time.isoformat()}",
        f"REGISTRATION:{registration}",
        f"STAY_TYPE:{duration_label}",
        f"DEPOSIT:{money(deposit_amount)}",
        f"DEPOSIT_PAYMENT:{deposit_payment}",
        f"KELESTARIAN_PAYMENT:{kelestarian_payment}",
    ]
    remark = clean(form.get("remark"))
    if remark:
        flags.append(f"REMARK:{remark}")
    return Stay(
        bill_no=f"{identity}-BILL",
        registration_no=registration,
        folio_no=identity,
        guest_name=clean(form.get("guest_name")) or f"Room {room}",
        room_no=room,
        departure_date=check_out,
        check_in_date=check_in,
        check_out_date=check_out,
        number_of_pax=1,
        number_of_nights=nights,
        rate_type=rate_type,
        tariff_before_tax=amount,
        tax_amount=Decimal("0"),
        food_amount=Decimal("0"),
        bar_amount=Decimal("0"),
        laundry_amount=Decimal("0"),
        call_amount=Decimal("0"),
        misc_amount=Decimal("0"),
        total_amount=amount,
        amount_paid=amount,
        payment_method=payment_method,
        flags=flags,
    )


def historical_summary(stays: list[Stay], fee_rate: Decimal, selected_year: int | None = None) -> dict:
    historical = [stay for stay in stays if stay.check_out_date and stay.check_out_date < KELESTARIAN_START]
    if selected_year:
        historical = [stay for stay in historical if stay.check_out_date.year == selected_year]
    by_year = defaultdict(lambda: {"stays": 0, "nights": 0, "revenue": Decimal("0")})
    by_month = defaultdict(lambda: {"stays": 0, "nights": 0, "revenue": Decimal("0")})
    payment = defaultdict(lambda: {"count": 0, "total_amount": Decimal("0")})
    guests = defaultdict(lambda: {"guest_name": "", "stays": 0, "nights": 0, "total_amount": Decimal("0"), "first": None, "latest": None})
    issues = []
    for stay in historical:
        year = str(stay.check_out_date.year)
        month = stay.check_out_date.strftime("%Y-%m")
        nights = stay.number_of_nights or 0
        by_year[year]["stays"] += 1
        by_year[year]["nights"] += nights
        by_year[year]["revenue"] += stay.total_amount
        by_month[month]["stays"] += 1
        by_month[month]["nights"] += nights
        by_month[month]["revenue"] += stay.total_amount
        pay = stay.payment_method or "Unknown"
        payment[pay]["count"] += 1
        payment[pay]["total_amount"] += stay.amount_paid
        guest_key = verification_value(stay.guest_name) or "UNKNOWN"
        guest = guests[guest_key]
        guest["guest_name"] = stay.guest_name or "Unknown"
        guest["stays"] += 1
        guest["nights"] += nights
        guest["total_amount"] += stay.total_amount
        guest["first"] = min(filter(None, [guest["first"], stay.check_in_date])) if guest["first"] else stay.check_in_date
        guest["latest"] = max(filter(None, [guest["latest"], stay.check_out_date])) if guest["latest"] else stay.check_out_date
        if stay.flags:
            issues.append({"bill_no": stay.bill_no, "guest_name": stay.guest_name, "room_no": stay.room_no, "flags": stay.flags})
    monthly_values = sorted(by_month.items())
    yearly_values = sorted(by_year.items())
    monthly_average_revenue = sum((item["revenue"] for _, item in monthly_values), Decimal("0")) / Decimal(len(monthly_values)) if monthly_values else Decimal("0")
    forecast_source_years = [year for year in ("2023", "2024", "2025") if year in by_year]
    forecast_source_nights = sum((by_year[year]["nights"] for year in forecast_source_years), 0)
    forecast_years = len(forecast_source_years) or 1
    forecast_nights = Decimal(forecast_source_nights) / Decimal(forecast_years)
    top_guests = sorted(guests.values(), key=lambda item: item["total_amount"], reverse=True)[:20]
    return {
        "selected_year": str(selected_year or ""),
        "historical_stays": len(historical),
        "first_year": yearly_values[0][0] if yearly_values else "",
        "last_year": yearly_values[-1][0] if yearly_values else "",
        "total_revenue": money(sum((stay.total_amount for stay in historical), Decimal("0"))),
        "total_nights": sum((stay.number_of_nights or 0) for stay in historical),
        "monthly_average_revenue": money(monthly_average_revenue),
        "forecast_2026_kelestarian": money(forecast_nights * fee_rate),
        "yearly_revenue": [
            {"year": year, "stays": item["stays"], "nights": item["nights"], "revenue": money(item["revenue"])}
            for year, item in yearly_values
        ],
        "monthly_revenue": [
            {"month": month, "stays": item["stays"], "nights": item["nights"], "revenue": money(item["revenue"])}
            for month, item in monthly_values
        ],
        "payment_categories": [
            {"payment_method": key, "count": item["count"], "total_amount": money(item["total_amount"])}
            for key, item in sorted(payment.items(), key=lambda row: row[1]["total_amount"], reverse=True)
        ],
        "top_guests": [
            {
                "guest_name": item["guest_name"],
                "stays": item["stays"],
                "nights": item["nights"],
                "total_amount": money(item["total_amount"]),
                "first": item["first"].isoformat() if item["first"] else "",
                "latest": item["latest"].isoformat() if item["latest"] else "",
            }
            for item in top_guests
        ],
        "data_quality_issues": issues[:200],
    }


CREST_PATH = os.path.join(os.path.dirname(__file__), "assets", "selangor-crest.jpeg")
GREY = colors.HexColor("#D9D9D9")
BODY_FONT = "Helvetica"
BOLD_FONT = "Helvetica-Bold"


def _text(c, x, y, value, size=11, bold=False, align="left"):
    value = clean(value)
    c.setFillColor(colors.black)
    c.setFont(BOLD_FONT if bold else BODY_FONT, size)
    if align == "center":
        c.drawCentredString(x, y, value)
    elif align == "right":
        c.drawRightString(x, y, value)
    else:
        c.drawString(x, y, value)


def _fit_size(value, width, size, bold=False, minimum=7.5):
    value = clean(value)
    font = BOLD_FONT if bold else BODY_FONT
    while size > minimum and stringWidth(value, font, size) > width:
        size -= 0.25
    return size


def _wrap_lines(value, width, size=10.5, bold=False, max_lines=2):
    words = clean(value).split()
    font = BOLD_FONT if bold else BODY_FONT
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if current and stringWidth(candidate, font, size) > width:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines[:max_lines]


def _outer(c, label, x, right, bottom, top=743):
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.55)
    c.rect(x, bottom, right - x, top - bottom, stroke=1, fill=0)
    _text(c, right - 22, top + 3, label, 12, True, "right")


def _crest(c, x=281.5, y=645, width=49, height=72):
    if os.path.exists(CREST_PATH):
        c.drawImage(ImageReader(CREST_PATH), x, y, width, height, preserveAspectRatio=True, mask="auto")


def _grid(c, x, top, widths, row_heights, fills=None):
    y = top
    total_width = sum(widths)
    for row_no, height in enumerate(row_heights):
        y -= height
        if fills and fills[row_no]:
            c.setFillColor(fills[row_no])
            c.rect(x, y, total_width, height, stroke=0, fill=1)
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.45)
        c.rect(x, y, total_width, height, stroke=1, fill=0)
        cursor = x
        for width in widths[:-1]:
            cursor += width
            c.line(cursor, y, cursor, y + height)
    return y


def _cell_text(c, x, y, width, height, value, size=10, bold=False, align="center"):
    size = _fit_size(value, width - 7, size, bold)
    baseline = y + (height - size) / 2 + 1
    if align == "left":
        _text(c, x + 4, baseline, value, size, bold)
    else:
        _text(c, x + width / 2, baseline, value, size, bold, "center")


def _b_fields(c, labels, values):
    label_x = 62.5
    colon_x = 280.5
    box_x = 288.5
    box_width = 266.5
    top = 570
    heights = [20.5, 20.5, 20.5, 41, 20.5, 20.5, 20.5, 20.5]
    y = top
    for index, (label, value, height) in enumerate(zip(labels, values, heights)):
        y -= height
        label_size = _fit_size(label, colon_x - label_x - 8, 11.5)
        _text(c, label_x, y + (height - label_size) / 2 + 1, label, label_size)
        _text(c, colon_x, y + (height - 11) / 2 + 1, ":", 11)
        c.rect(box_x, y, box_width, height, stroke=1, fill=0)
        if index == 3:
            lines = _wrap_lines(value, box_width - 12, 10.5, max_lines=2)
            for line_no, line in enumerate(lines):
                _text(c, box_x + 6, y + height - 14 - line_no * 14, line, 10.5)
        else:
            size = _fit_size(value, box_width - 12, 10.5)
            _text(c, box_x + 6, y + (height - size) / 2 + 1, value, size)
    return y


B_WIDTHS = [87, 142.5, 126, 143]


def _b_table(c, top, rows, include_header=False, include_total=None):
    rendered = []
    fills = []
    if include_header:
        rendered.append(["Tarikh", "Jumlah Bilik (Unit)", "Bilangan Malam", "Jumlah Kutipan (RM)"])
        fills.append(GREY)
    rendered.extend(rows)
    fills.extend([None] * len(rows))
    if include_total is not None:
        rendered.append(include_total)
        fills.append(GREY)
    heights = [17.5] * len(rendered)
    bottom = _grid(c, 56.5, top, B_WIDTHS, heights, fills)
    y = top
    for row_no, row in enumerate(rendered):
        y -= heights[row_no]
        cursor = 56.5
        for col_no, value in enumerate(row):
            _cell_text(c, cursor, y, B_WIDTHS[col_no], heights[row_no], value, 10.5, bool(fills[row_no]))
            cursor += B_WIDTHS[col_no]
    return bottom


def _b_confirmation(c, total_fee, table_bottom):
    heading_y = table_bottom - 27
    _text(c, 56.5, heading_y, "C. PENGESAHAN OLEH PENGUSAHA PREMIS PENGINAPAN", 12, True)
    _text(c, 56.5, heading_y - 26, "Saya mengesahkan bahawa maklumat ini adalah BENAR dan TEPAT berdasarkan rekod", 11)
    _text(c, 56.5, heading_y - 41, "kutipan SEBENAR bagi bulan yang dilaporkan.", 11)
    _text(c, 56.5, heading_y - 71, f"Jumlah Kutipan Bulanan (RM) : {money(total_fee)}", 11.5, True)
    _text(c, 56.5, heading_y - 88, f"Tarikh : {datetime.now().strftime('%d/%m/%Y')}", 11.5, True)
    _text(c, 56.5, heading_y - 157, "................................................", 11, True)
    _text(c, 56.5, heading_y - 172, "Cop Rasmi & Tandatangan:", 11.5, True)
    divider = heading_y - 188
    c.line(50.5, divider, 561, divider)
    _text(c, 56.5, divider - 22, "D. UNTUK KEGUNAAN PEJABAT PBT SAHAJA", 12, True)
    _text(c, 56.5, divider - 50, "(SEMAKAN & PENGESAHAN PBT)", 12, True)
    pbt = ["Tarikh Diterima", "Ulasan / Catatan", "Jumlah Kutipan Bulanan (RM)", "Caj lewat (RM) (Sekiranya ada)"]
    for index, label in enumerate(pbt):
        y = divider - 83 - index * 17
        _text(c, 62, y, label, 11)
        _text(c, 240, y, ": ................................................", 11)
    _text(c, 56.5, divider - 202, "................................................", 11, True)
    _text(c, 56.5, divider - 217, "Cop Rasmi & Tandatangan:", 11.5, True)


def report_filename(kind: str, stays: list[Stay], report_start: date | None = None, report_end: date | None = None) -> str:
    if report_start and report_end:
        if kind == "b" and report_start.year == report_end.year and report_start.month == report_end.month:
            return f"Lampiran B {month_name[report_start.month]} {report_start.year}.pdf"
        if kind == "c":
            start_label = f"{report_start.day:02d} {month_name[report_start.month]} {report_start.year}"
            end_label = f"{report_end.day:02d} {month_name[report_end.month]} {report_end.year}"
            period = start_label if report_start == report_end else f"{start_label} to {end_label}"
            return f"Lampiran C {period}.pdf"
    check_in_dates = sorted({stay.check_in_date for stay in stays if stay.check_in_date})
    if kind == "b":
        months = sorted({(day.year, day.month) for day in check_in_dates})
        labels = [f"{month_name[month]} {year}" for year, month in months]
        period = labels[0] if len(labels) == 1 else f"{labels[0]} to {labels[-1]}"
        return f"Lampiran B {period}.pdf"
    labels = [f"{day.day:02d} {month_name[day.month]} {day.year}" for day in check_in_dates]
    period = labels[0] if len(labels) == 1 else f"{labels[0]} to {labels[-1]}"
    return f"Lampiran C {period}.pdf"


def stay_overlaps_period(stay: Stay, start: date, end: date) -> bool:
    if not stay.check_in_date or not stay.number_of_nights:
        return False
    if stay.check_in_date < KELESTARIAN_START or stay.check_in_date > KELESTARIAN_END:
        return False
    stay_last_night = stay.check_in_date + timedelta(days=stay.number_of_nights - 1)
    charge_start = max(start, KELESTARIAN_START)
    charge_end = min(end, KELESTARIAN_END)
    return charge_start <= charge_end and stay.check_in_date <= charge_end and stay_last_night >= charge_start


def filter_stays_for_report(stays: list[Stay], kind: str, report_month="", report_start="", report_end="") -> list[Stay]:
    if report_month:
        if not re.fullmatch(r"\d{4}-\d{2}", report_month):
            raise ValueError(f"Please select a valid month for Lampiran {kind.upper()}.")
        year, month = (int(part) for part in report_month.split("-"))
        start = date(year, month, 1)
        end = date(year, month, monthrange(year, month)[1])
        selected = [stay for stay in stays if stay_overlaps_period(stay, start, end)]
    elif kind == "c" and (report_start or report_end):
        try:
            start = date.fromisoformat(report_start or report_end)
            end = date.fromisoformat(report_end or report_start)
        except ValueError as exc:
            raise ValueError("Please select valid dates for Lampiran C.") from exc
        if start > end:
            raise ValueError("Lampiran C start date must be before or equal to the end date.")
        selected = [stay for stay in stays if stay_overlaps_period(stay, start, end)]
    else:
        selected = [stay for stay in stays if stay_overlaps_period(stay, KELESTARIAN_START, KELESTARIAN_END)]
    if not selected:
        label = "month" if kind == "b" else "date range"
        raise ValueError(f"No guest check-ins were found for the selected {label}.")
    return selected


def report_sales_rows(stays: list[Stay], fee_rate: Decimal, report_start: date | None = None, report_end: date | None = None) -> list[dict]:
    rows = expanded_sales(stays, fee_rate)
    if report_start:
        rows = [row for row in rows if row["date"] >= report_start.isoformat()]
    if report_end:
        rows = [row for row in rows if row["date"] <= report_end.isoformat()]
    return rows


def form_b_pdf(stays: list[Stay], settings: dict, fee_rate: Decimal, report_start: date | None = None, report_end: date | None = None) -> bytes:
    check_in_rows = [row for row in report_sales_rows(stays, fee_rate, report_start, report_end) if not row["is_paid_continuation"]]
    months = defaultdict(list)
    for row in check_in_rows:
        day = date.fromisoformat(row["date"])
        months[(day.year, day.month)].append(row)
    out = io.BytesIO()
    c = canvas.Canvas(out, pagesize=LETTER)
    for month_index, ((year, month), month_rows) in enumerate(sorted(months.items())):
        if month_index:
            c.showPage()
        daily = defaultdict(list)
        for row in month_rows:
            daily[date.fromisoformat(row["date"]).day].append(row)
        values = [
            settings.get("premise_name", ""), settings.get("license_no", ""),
            settings.get("certificate_no", ""), settings.get("address", ""),
            settings.get("contact_name", ""), settings.get("contact", ""),
            settings.get("category_code", ""), f"{month_name[month]} {year}",
        ]
        labels = ["Nama Premis Penginapan", "No. Lesen Perniagaan (PBT)", "No. Siri Sijil", "Alamat Premis Penginapan", "Nama wakil premis untuk dihubungi", "No. Telefon / Emel", "Kod Kategori Premis Penginapan", "Bulan & Tahun Pelaporan"]
        _outer(c, "LAMPIRAN B", 50.5, 561, 14.5)
        _crest(c)
        _text(c, LETTER[0] / 2, 620, "LAPORAN PENYATA KUTIPAN FI KELESTARIAN NEGERI SELANGOR (BULANAN)", 12, True, "center")
        _text(c, 56.5, 589, "A. MAKLUMAT PREMIS PENGINAPAN", 12, True)
        _b_fields(c, labels, values)
        _text(c, 56.5, 365, "B. MAKLUMAT KUTIPAN FI KELESTARIAN", 12, True)
        day_rows = []
        total_rooms = total_nights = 0
        total_fee = Decimal("0")
        for day_no in range(1, monthrange(year, month)[1] + 1):
            collected = daily.get(day_no, [])
            rooms = len(collected)
            nights = sum(item["nights"] for item in collected)
            fee = sum((dec(item["kelestarian"]) for item in collected), Decimal("0"))
            total_rooms += rooms
            total_nights += nights
            total_fee += fee
            day_rows.append([f"{day_no}/{month:02d}/{year}", str(rooms) if rooms else "", str(nights) if nights else "", money(fee) if fee else ""])
        _b_table(c, 347.5, day_rows[:18], include_header=True)
        c.showPage()
        _outer(c, "LAMPIRAN B", 50.5, 561, 29)
        bottom = _b_table(c, 743, day_rows[18:], include_total=["Jumlah", str(total_rooms), str(total_nights), money(total_fee)])
        _b_confirmation(c, total_fee, bottom)
    c.save()
    return out.getvalue()


C_WIDTHS = [29.5, 276.5, 85, 71, 107.5]
C_HEADERS = ["Bil.", "Nama", "Jumlah Bilik\n(Unit)", "Bilangan\nMalam", "Jumlah Fi\nKelestarian (RM)"]
C_MANUAL_ROWS = 5


def _c_header_cells(c, x, y, height=40):
    cursor = x
    for width, label in zip(C_WIDTHS, C_HEADERS):
        lines = label.split("\n")
        baseline = y + height / 2 + (len(lines) - 1) * 6 - 4
        for line_no, line in enumerate(lines):
            size = _fit_size(line, width - 6, 11.5, True, 8.5)
            _text(c, cursor + width / 2, baseline - line_no * 13, line, size, True, "center")
        cursor += width


def _c_rows(c, top, items, capacity, start_number, include_header=False):
    header_height = 40 if include_header else 0
    row_height = 17.5
    padded = items + [None] * max(0, capacity - len(items))
    fills = ([GREY] if include_header else []) + [None] * len(padded)
    heights = ([header_height] if include_header else []) + [row_height] * len(padded)
    bottom = _grid(c, 21, top, C_WIDTHS, heights, fills)
    y = top
    if include_header:
        y -= header_height
        _c_header_cells(c, 21, y, header_height)
    for local_no, item in enumerate(padded):
        y -= row_height
        if item is None:
            continue
        values = [str(start_number + local_no), item["guest_name"], "1", str(item["nights"]), item["kelestarian"]]
        cursor = 21
        for col_no, value in enumerate(values):
            _cell_text(c, cursor, y, C_WIDTHS[col_no], row_height, value, 10.5, False, "left" if col_no == 1 else "center")
            cursor += C_WIDTHS[col_no]
    return bottom


def _c_first_page(c, settings, day_label, items):
    _outer(c, "LAMPIRAN C", 15.5, 596, 35.5)
    _crest(c, 281.5, 664, 49, 72)
    _text(c, LETTER[0] / 2, 641, "LAPORAN TRANSAKSI PENGGUNAAN BILIK (HARIAN)", 12, True, "center")
    labels = ["Nama Premis Penginapan", "No. Rujukan Lesen (PBT)", "No. Siri Sijil", "Tarikh Hari Daftar Masuk (Check-In)"]
    values = [settings.get("premise_name", ""), settings.get("license_no", ""), settings.get("certificate_no", ""), day_label]
    box_tops = [635, 606, 576.5, 547.5]
    for label, value, box_top in zip(labels, values, box_tops):
        size = _fit_size(label, 215, 12)
        _text(c, 40.5, box_top - 12, label, size)
        _text(c, 263, box_top - 12, ":", 12)
        c.rect(270.5, box_top - 15, 306.5, 15, stroke=1, fill=0)
        value_size = _fit_size(value, 294, 10.5)
        _text(c, 276.5, box_top - 12, value, value_size)
    _c_rows(c, 513.5, items, 25, 1, include_header=True)


def _c_confirmation(c, table_bottom, rows_for_day):
    total_rooms = len(rows_for_day)
    total_nights = sum(row["nights"] for row in rows_for_day)
    total_fee = sum((dec(row["kelestarian"]) for row in rows_for_day), Decimal("0"))
    height = 21.5
    c.setFillColor(GREY)
    c.rect(21, table_bottom - height, sum(C_WIDTHS), height, stroke=0, fill=1)
    c.setStrokeColor(colors.black)
    c.rect(21, table_bottom - height, sum(C_WIDTHS), height, stroke=1, fill=0)
    cursor = 21 + sum(C_WIDTHS[:2])
    for width in C_WIDTHS[2:]:
        c.line(cursor, table_bottom - height, cursor, table_bottom)
        cursor += width
    _text(c, 21 + sum(C_WIDTHS[:2]) / 2, table_bottom - 15, "Jumlah Keseluruhan (Unit @ RM)", 11.5, True, "center")
    cursor = 21 + sum(C_WIDTHS[:2])
    for width, value in zip(C_WIDTHS[2:], [str(total_rooms), str(total_nights), money(total_fee)]):
        _cell_text(c, cursor, table_bottom - height, width, height, value, 10.5, True)
        cursor += width
    box_top = table_bottom - height - 11
    c.line(15.5, box_top, 596, box_top)
    _text(c, 21, box_top - 34, "PENGESAHAN OLEH PENGUSAHA PREMIS PENGINAPAN", 12, True)
    _text(c, 21, box_top - 57, "Saya mengesahkan bahawa maklumat ini adalah BENAR dan TEPAT berdasarkan rekod kutipan", 11)
    _text(c, 21, box_top - 72, "SEBENAR bagi hari yang dilaporkan.", 11, True)
    _text(c, 21, box_top - 101, f"Jumlah Kutipan Harian (RM) : {money(total_fee)}", 11.5, True)
    _text(c, 21, box_top - 118, f"Tarikh : {datetime.now().strftime('%d/%m/%Y')}", 11.5, True)
    _text(c, 21, box_top - 185, "................................................", 11, True)
    _text(c, 21, box_top - 200, "Cop Rasmi & Tandatangan:", 11.5, True)


def form_c_pdf(stays: list[Stay], settings: dict, fee_rate: Decimal, report_start: date | None = None, report_end: date | None = None) -> bytes:
    grouped = defaultdict(list)
    for row in report_sales_rows(stays, fee_rate, report_start, report_end):
        if not row["is_paid_continuation"]:
            grouped[row["date"]].append(row)
    out = io.BytesIO()
    c = canvas.Canvas(out, pagesize=LETTER)
    first_report = True
    for day_key in sorted(grouped):
        if not first_report:
            c.showPage()
        first_report = False
        rows_for_day = grouped[day_key]
        day_label = date.fromisoformat(day_key).strftime("%d/%m/%Y")
        _c_first_page(c, settings, day_label, rows_for_day[:25])
        remaining = rows_for_day[25:]
        row_number = 26
        while len(remaining) > 16:
            c.showPage()
            _outer(c, "LAMPIRAN C", 15.5, 596, 35.5)
            chunk_size = min(35, len(remaining) - 16)
            chunk = remaining[:chunk_size]
            _c_rows(c, 743, chunk, len(chunk), row_number, include_header=True)
            row_number += len(chunk)
            remaining = remaining[len(chunk):]
        c.showPage()
        _outer(c, "LAMPIRAN C", 15.5, 596, 134)
        bottom = _c_rows(c, 743, remaining, len(remaining) + C_MANUAL_ROWS, row_number, include_header=False)
        _c_confirmation(c, bottom, rows_for_day)
    c.save()
    return out.getvalue()


PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>KL Guest Hotel Sales</title>
<style>
select{width:100%;border:1px solid #c9d5e3;border-radius:7px;padding:11px 12px;font:inherit;background:#fff}.toolbar select{max-width:180px}
textarea{width:100%;min-height:94px;border:1px solid #c9d5e3;border-radius:7px;padding:11px 12px;font:inherit;resize:vertical}
:root{font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",Arial,sans-serif;background:#f2f6fb;color:#071a36}*{box-sizing:border-box}html,body{height:100%;overflow:hidden}body{margin:0}.app{height:100vh;display:grid;grid-template-columns:278px minmax(0,1fr);overflow:hidden}.sidebar{background:#12233b;color:#fff;padding:28px 18px;display:flex;flex-direction:column;overflow:hidden}.brand{display:flex;gap:14px;align-items:center;margin-bottom:30px}.logo{width:48px;height:48px;border-radius:9px;background:#fff;color:#0d3265;display:grid;place-items:center;font-weight:900}.brand span{display:block;color:#b5cae8;font-size:13px}.nav-label{font-size:12px;letter-spacing:.12em;text-transform:uppercase;color:#8aa6c8;font-weight:900}.workspace-select{width:100%;margin-top:8px;margin-bottom:28px;border:1px solid rgba(255,255,255,.82);border-radius:10px;background:#12233b;color:#fff;padding:14px 18px;font-size:18px;font-weight:800}.workspace-select:focus{outline:3px solid rgba(138,166,200,.35)}.nav{display:grid;gap:8px;margin-top:12px}.nav button{background:transparent;color:#b8d0ee;text-align:left;border:0;padding:14px;border-radius:8px;font-size:17px;cursor:pointer}.nav button.active{background:#1d3a60;color:#fff}.account{border-top:1px solid #314966;margin-top:auto;padding-top:28px;display:flex;gap:12px;align-items:center}.avatar{width:42px;height:42px;border-radius:999px;background:#3478d4;display:grid;place-items:center;font-weight:900}.main{min-width:0;overflow-y:auto;padding:28px 34px 42px}.top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:22px}.eyebrow{margin:0 0 8px;color:#7587a0;text-transform:uppercase;letter-spacing:.14em;font-size:13px;font-weight:900}h1{font-size:32px;margin:0}h2{font-size:20px;margin:0}.badge{background:#e8f1fd;color:#1c5a9d;border-radius:999px;padding:9px 16px;font-weight:900;font-size:13px}.notice{background:#e8f2ff;border:1px solid #bdd6fb;color:#0b4f9a;border-radius:9px;padding:16px 18px;margin-bottom:18px}.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:18px}.final-row{display:grid;grid-template-columns:minmax(260px,390px);gap:14px;margin:-4px 0 28px}.metric,.panel{background:#fff;border:1px solid #d7e0eb;border-radius:10px;box-shadow:0 8px 24px rgba(10,31,68,.04)}.metric{padding:16px}.metric.final{background:#168244;border-color:#168244;color:#fff;box-shadow:0 12px 28px rgba(22,130,68,.22)}.metric.final span{color:#d9f7e6}.metric.final strong{color:#fff}.metric span{display:block;color:#71839c;font-size:13px}.metric strong{display:block;margin-top:4px;font-size:22px}.panel{padding:22px;margin-bottom:18px}.cashflow-transfer{border-top:1px solid #e3ebf5;margin-top:8px;padding-top:20px}.cashflow-transfer .manual-row{grid-template-columns:repeat(4,minmax(150px,1fr)) auto}.drop{height:118px;border:1.5px dashed #9cb9dc;border-radius:9px;background:#f8fbff;display:grid;place-items:center;text-align:center;color:#1b5fab;font-weight:900;cursor:pointer}.drop.drag{background:#e8f2ff;border-color:#1b5fab}.drop small{display:block;color:#71839c;font-weight:600;margin-top:5px}.drop input{display:none}.toolbar{display:flex;justify-content:space-between;align-items:center;gap:12px;margin:14px 0;flex-wrap:wrap}.toolbar input{max-width:320px}.filters,.report-controls{display:flex;gap:8px;align-items:end;flex-wrap:wrap;margin:0 0 14px}.filters button{border:1px solid #c7d7ea;background:#fff;color:#1a4f8d;border-radius:7px;padding:9px 12px;font-weight:900;cursor:pointer}.filters button.active{background:#1d5fa7;color:#fff;border-color:#1d5fa7}.filters input{width:150px}.report-controls label{min-width:190px}.settings-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}.manual-row{display:grid;grid-template-columns:90px 138px 130px 94px 138px 132px 132px 132px auto;gap:10px;align-items:end}.manual-row label{min-width:0}.manual-row button{height:44px}.manual-checkin-grid{display:grid;grid-template-columns:repeat(4,minmax(170px,1fr));gap:16px;align-items:end}.manual-checkin-grid .wide-field{grid-column:span 2}.manual-checkin-actions{display:flex;justify-content:flex-end;align-items:end}.manual-checkin-actions .primary{height:46px;min-width:140px}.manual-remark-row{margin-top:16px;display:grid;grid-template-columns:minmax(260px,1fr);max-width:760px}.manual-remark-row input{min-height:46px}.rate-list{display:grid;gap:10px}.rate-list div{display:grid;grid-template-columns:130px 1fr;gap:12px;align-items:center}.room-help{color:#5e728c;font-size:13px;line-height:1.55}label{display:grid;gap:7px;font-size:13px;font-weight:800;color:#2f4360}input{width:100%;border:1px solid #c9d5e3;border-radius:7px;padding:11px 12px;font:inherit}button.primary{border:0;border-radius:8px;background:#10233e;color:#fff;font-weight:900;padding:12px 18px;cursor:pointer}button.primary:disabled{opacity:.45;cursor:not-allowed}button.danger{border:1px solid #f0b8b8;background:#fff5f5;color:#9b1c1c;border-radius:7px;padding:7px 10px;font-weight:900;cursor:pointer}.view{display:none}.view.active{display:block}.table-wrap{overflow:auto;border:1px solid #dde6f1;border-radius:8px;background:#fff}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:10px 12px;border-bottom:1px solid #e7edf5;text-align:left;white-space:nowrap}th{background:#f5f8fc;color:#516985;font-size:12px;text-transform:uppercase;letter-spacing:.05em}.day-row td{background:#edf4fc;color:#153a63;font-weight:900;font-size:14px}.multi td{background:#fff7dd}.paid-continuation td{background:#eaf8ef;color:#16633a}.paid-pill{display:inline-block;background:#dff5e8;color:#16633a;border:1px solid #a9dfbf;border-radius:999px;padding:4px 9px;font-weight:900}.review-reason{color:#9b2c2c;font-weight:900}.empty{color:#71839c;padding:22px;border:1px dashed #c9d8ea;border-radius:8px;background:#f9fbfe}.report-grid{display:grid;grid-template-columns:1fr;gap:18px}.wide{grid-column:span 2}@media(max-width:1500px){.manual-row{grid-template-columns:repeat(5,1fr)}.manual-checkin-grid{grid-template-columns:repeat(3,minmax(170px,1fr))}.cashflow-transfer .manual-row{grid-template-columns:repeat(4,1fr)}}@media(max-width:1050px){.app{grid-template-columns:1fr}.sidebar{display:none}.cards,.settings-grid,.manual-row,.manual-checkin-grid,.cashflow-transfer .manual-row{grid-template-columns:1fr}.final-row{grid-template-columns:1fr}.manual-checkin-grid .wide-field{grid-column:auto}.manual-checkin-actions{justify-content:stretch}.manual-checkin-actions .primary{width:100%}.manual-remark-row{max-width:none}.main{padding:20px}.top{flex-direction:column;gap:12px}}
.report-card-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px}.report-card{border:1px solid #d7e0eb;border-radius:9px;background:#fff;padding:16px;box-shadow:0 8px 22px rgba(10,31,68,.04)}.report-card h3{margin:0 0 12px;font-size:17px}.report-card dl{margin:0;display:grid;gap:9px}.report-card div{display:flex;justify-content:space-between;gap:12px;border-top:1px solid #edf2f8;padding-top:9px}.report-card div:first-child{border-top:0;padding-top:0}.report-card dt{color:#667b96;font-size:12px;text-transform:uppercase;letter-spacing:.05em;font-weight:900}.report-card dd{margin:0;font-weight:900;text-align:right}@media(max-width:1050px){.report-card-grid{grid-template-columns:1fr}}
#analytics>.cards{display:none}#analytics .table-wrap{border:0;background:transparent;overflow:visible}#analytics table,#analytics thead,#analytics tbody{display:block}#analytics thead{display:none}#analytics tbody{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px}#analytics tr{display:block;border:1px solid #d7e0eb;border-radius:9px;background:#fff;padding:16px;box-shadow:0 8px 22px rgba(10,31,68,.04)}#analytics td{display:flex;justify-content:space-between;gap:12px;border:0;border-top:1px solid #edf2f8;padding:9px 0 0;white-space:normal}#analytics td:first-child{display:block;border-top:0;padding-top:0;margin-bottom:4px;font-size:17px;font-weight:900}#analytics td:nth-child(2)::before{content:"Transactions";color:#667b96;font-size:12px;text-transform:uppercase;letter-spacing:.05em;font-weight:900}#analytics td:nth-child(3)::before{content:"Amount";color:#667b96;font-size:12px;text-transform:uppercase;letter-spacing:.05em;font-weight:900}#analytics td:nth-child(4)::before{content:"Lestari Fee";color:#667b96;font-size:12px;text-transform:uppercase;letter-spacing:.05em;font-weight:900}
</style>
</head>
<body>
<main class="app">
  <aside class="sidebar">
    <div class="brand"><div class="logo">KL</div><div><strong>KL Guest Hotel</strong><span>Sales & Lampiran</span></div></div>
    <div class="nav-label">Workspace</div>
    <select id="workspaceSelect" class="workspace-select" aria-label="Workspace">
      <option value="sales">Sales</option>
      <option value="dashboard">Dashboard</option>
      <option value="lestari">Lestari</option>
    </select>
    <nav class="nav">
      <button data-workspace="sales" data-view="importSales" class="active">Import Data</button>
      <button data-workspace="sales" data-view="manual">Manual Check-In</button>
      <button data-workspace="dashboard" data-view="dashboard">Dashboard</button>
      <button data-workspace="dashboard" data-view="dailySales">Daily Sales</button>
      <button data-workspace="dashboard" data-view="historical">Historical</button>
      <button data-workspace="dashboard" data-view="review">Manual Review</button>
      <button data-workspace="dashboard" data-view="settings">Settings</button>
      <button data-workspace="lestari" data-view="b">Lampiran B</button>
      <button data-workspace="lestari" data-view="c">Lampiran C</button>
    </nav>
    <div class="account"><div class="avatar">A</div><div><strong>Admin</strong><br><span style="color:#a6bddb">@admin</span></div></div>
  </aside>
  <section class="main">
    <header class="top"><div><p class="eyebrow">Sales Tracking Workspace</p><h1 id="pageTitle">Import Data</h1></div><div class="badge">ADMIN</div></header>
    <section class="notice" id="notice">Drag in the Sales Bill Register Excel file. The dashboard will remember every guest stay and expand multi-night bills by day.</section>
    <section class="cards">
      <span id="mStays" hidden>0</span>
      <span id="mRows" hidden>0</span>
      <article class="metric"><span>Total sales</span><strong id="mSales">RM 0.00</strong></article>
      <article class="metric"><span>Kelestarian</span><strong id="mFee">RM 0.00</strong></article>
    </section>
    <section id="dashboard" class="view">
      <div class="panel"><h2>Sales reports</h2><br><div id="analytics" class="empty">Import Excel first.</div></div>
      <div class="panel"><div class="toolbar"><h2>Daily sales ledger</h2><input id="search" placeholder="Search guest or room"></div><div class="filters"><button data-range="today">Today</button><button data-range="last7">Last 7 days</button><button data-range="month" class="active">This month</button><button data-range="custom">Custom date</button><input id="startDate" type="date"><input id="endDate" type="date"></div><div id="ledger" class="empty">No imported check-ins yet.</div></div>
    </section>
    <section id="importSales" class="view active">
      <div class="panel"><div class="toolbar"><h2>Import data</h2></div><label class="drop" id="drop"><span id="dropText">Choose or drop Sales Bill Register Excel<small>.xls or .xlsx</small></span><input id="file" type="file" accept=".xls,.xlsx"></label></div>
    </section>
    <section id="manual" class="view">
      <div class="panel"><div class="toolbar"><h2>Manual daily check-in</h2></div><div class="manual-checkin-grid">
        <label>Check-in date<input id="manualCheckIn" class="manual" name="check_in_date" type="date"></label>
        <label>Room<input id="manualRoom" class="manual" name="room" placeholder="305"></label>
        <label>Registration<select id="manualRegistration" class="manual" name="registration"><option>walk in</option><option>online</option><option>extension</option></select></label>
        <label>Stay type<select id="manualStayType" class="manual" name="stay_type"><option value="night">Night</option><option value="3_hour">Day use</option></select></label>
        <label>Stay duration<input id="manualDuration" class="manual" name="stay_duration" type="number" min="1" step="1" value="1"></label>
        <label>Room payment (RM)<input id="manualAmount" class="manual" name="amount_paid" type="number" step="0.01"></label>
        <label class="wide-field">Payment method<select id="manualPayment" class="manual" name="payment_method"><option value="cash">Cash</option><option value="qr">QR</option><option value="card">Card</option><option value="online">Online Booking</option><option value="transfer">Bank Transfer</option></select></label>
        <label>Kelestarian payment<select id="manualKelestarianPayment" class="manual" name="kelestarian_payment_method"><option value="cash">Cash</option><option value="qr">QR</option><option value="card">Card</option><option value="online">Online Booking</option><option value="transfer">Bank Transfer</option></select></label>
        <label>Deposit payment<select id="manualDepositPayment" class="manual" name="deposit_payment_method"><option value="cash">Cash</option><option value="qr">QR</option><option value="card">Card</option><option value="online">Online Booking</option><option value="transfer">Bank Transfer</option></select></label>
        <div class="manual-checkin-actions"><button class="primary" id="saveManual">Save</button></div>
        <input id="manualGuest" class="manual" name="guest_name" type="hidden" value="Walk-in guest">
        <input id="manualCheckOut" class="manual" name="check_out_date" type="hidden">
        <input id="manualDeposit" class="manual" name="deposit_amount" type="hidden" value="50.00">
      </div><div class="manual-remark-row"><label>Remark<input id="manualRemark" class="manual" name="remark" placeholder="Optional"></label></div><br><div id="manualPreview" class="empty">Key in the room to preview the default charges.</div></div>
      <div class="panel"><div class="toolbar"><h2>Transactions</h2><label>History date<input id="manualHistoryDate" type="date"></label></div><div id="manualHistory" class="empty">No manual check-ins saved yet.</div></div>
    </section>
    <section id="dailySales" class="view">
      <div class="panel"><div class="toolbar"><h2>Live daily dashboard</h2></div><div class="filters"><button data-daily-range="today" class="active">Today</button><button data-daily-range="custom">Custom date</button><input id="dailyStartDate" type="date"><input id="dailyEndDate" type="date"></div></div>
      <div class="panel"><h2>Cash / QR / Bank on hand</h2><br><div id="dailyCashflow" class="empty">No check-ins in this date range.</div><div class="cashflow-transfer"><div class="toolbar"><h2>Transfer between Cash / QR / Bank</h2></div><div class="manual-row">
        <label>From<select id="transferFrom"><option value="cash">Cash</option><option value="qr">QR</option><option value="transfer">Bank</option></select></label>
        <label>To<select id="transferTo"><option value="qr">QR</option><option value="cash">Cash</option><option value="transfer">Bank</option></select></label>
        <label>Amount (RM)<input id="transferAmount" type="number" step="0.01" min="0"></label>
        <label>Remark<input id="transferNote" placeholder="QR changed to cash"></label>
        <button class="primary" id="saveTransferMovement">Transfer</button>
      </div><br><div id="transferMovements" class="empty">No transfers for this date range.</div></div></div>
      <div class="panel"><div class="toolbar"><h2>Owner equity and drawings</h2><label>View owner<select id="ownerFilter"><option value="">All owners</option></select></label></div><div class="manual-row">
        <label>Owner<select id="ownerSelect"><option value="">Choose owner</option><option value="__new__">+ Add new owner</option></select></label>
        <label id="ownerNameWrap">New owner name<input id="ownerName" placeholder="Owner name"></label>
        <label>Type<select id="ownerType"><option value="equity">Owner Equity</option><option value="drawing">Drawings</option></select></label>
        <label>Method<select id="ownerMethod"><option value="cash">Cash</option><option value="qr">QR</option><option value="transfer">Bank</option></select></label>
        <label>Amount (RM)<input id="ownerAmount" type="number" step="0.01" min="0"></label>
        <label>Remark<input id="ownerNote" placeholder="Optional"></label>
        <button class="primary" id="saveOwnerMovement">Add</button>
      </div><br><div id="ownerMovements" class="empty">No owner equity or drawings history yet.</div></div>
      <div class="panel"><div class="toolbar"><h2>Live check-ins</h2></div><div id="dailySalesTable" class="empty">No check-ins in this date range.</div></div>
    </section>
    <section id="historical" class="view">
      <div class="panel"><div class="toolbar"><h2>Historical sales insights</h2><div class="toolbar"><select id="historicalYear"></select><button class="primary" id="loadHistorical">Load history</button></div></div><div id="historicalPanel" class="empty">Select a year to review sales history, payment trends, guest memory, forecasts, archive batches, and data quality.</div></div>
    </section>
    <section id="b" class="view">
      <div class="panel"><div class="toolbar"><h2>Laporan B preview</h2><button class="primary" id="downloadB" disabled>Download PDF</button></div><div class="report-controls"><label>Month to print<input id="reportMonthB" type="month"></label></div><div id="previewB" class="empty">Import Excel first.</div></div>
    </section>
    <section id="c" class="view">
      <div class="panel"><div class="toolbar"><h2>Laporan C preview</h2><button class="primary" id="downloadC" disabled>Download PDF</button></div><div class="report-controls"><label>Month to print<input id="reportMonthC" type="month"></label><label>First date<input id="reportStartC" type="date"></label><label>Last date<input id="reportEndC" type="date"></label></div><div id="previewC" class="empty">Import Excel first.</div></div>
    </section>
    <section id="review" class="view">
      <div class="panel"><div class="toolbar"><div><h2>Manual Review</h2><small>Folio and Bill Number mismatches are kept out of the guest database.</small></div><button class="primary" id="refreshReview">Refresh</button></div><div id="reviewQueue" class="empty">No records need manual review.</div></div>
    </section>
    <section id="settings" class="view">
      <div class="panel"><h2>Reporting settings</h2><br><div class="settings-grid">
        <label>Nama Premis Penginapan<input name="premise_name" class="setting" placeholder="KL Guest Hotel"></label>
        <label>No. Lesen Perniagaan (PBT)<input name="license_no" class="setting" placeholder="MBPJ-0000"></label>
        <label>No. Siri Sijil<input name="certificate_no" class="setting"></label>
        <label>Kod Kategori Premis<input name="category_code" class="setting" placeholder="Hotel 1-3 bintang"></label>
        <label>Wakil Untuk Dihubungi<input name="contact_name" class="setting"></label>
        <label>No. Telefon / Emel<input name="contact" class="setting" value="012-205-0039 / hueyjiunphang@gmail.com"></label>
        <label>Fi per bilik/malam (RM)<input name="fee_rate" class="setting" type="number" step="0.01" value="5.00"></label>
        <label class="wide">Alamat Premis Penginapan<input name="address" class="setting" value="8, Jalan AU 1a/4c, Taman Keramat Permai, 54200 Kuala Lumpur, Federal Territory of Kuala Lumpur"></label>
        <label>Twin room rate (RM)<input name="rate_twin" class="setting" type="number" step="0.01" value="108.00"></label>
        <label>Family room rate (RM)<input name="rate_family" class="setting" type="number" step="0.01" value="148.00"></label>
        <label>Family suite rate (RM)<input name="rate_suite" class="setting" type="number" step="0.01" value="168.00"></label>
        <label>Day use rate (RM)<input name="rate_day_use" class="setting" type="number" step="0.01" value="65.00"></label>
        <div class="wide room-help"><strong>Room groups</strong><br>Family suite: 101, 201, 301<br>Twin room: 105-102, 112-109, 205-202, 212-209, 305-302, 312-309<br>Family room: 106-108, 113-115, 206-208, 213-215, 306-308, 313-315<br>Online booking is always manual price.</div>
      </div></div>
    </section>
  </section>
</main>
<script>
const APP_VERSION="sales-report-cards-20260626";
if(localStorage.getItem("appVersion")!==APP_VERSION){["stays","sales","summary"].forEach(k=>localStorage.removeItem(k));localStorage.setItem("appVersion",APP_VERSION)}
let stays=JSON.parse(localStorage.getItem("stays")||"[]");
let sales=JSON.parse(localStorage.getItem("sales")||"[]");
let summary=JSON.parse(localStorage.getItem("summary")||"{}");
let rangeMode=localStorage.getItem("rangeMode")||"month";
let dailySalesMode=localStorage.getItem("dailySalesMode")||"today";
let ownerMovements=JSON.parse(localStorage.getItem("ownerMovements")||"[]");
let transferMovements=JSON.parse(localStorage.getItem("transferMovements")||"[]");
let reviewQueue=[];
const q=s=>document.querySelector(s), qa=s=>[...document.querySelectorAll(s)];
function settings(){let o={};qa(".setting").forEach(i=>o[i.name]=i.value);return o}
function asMoney(v){return Number(String(v||"0").replace(/,/g,""))||0}
function html(v){return String(v??"").replace(/[&<>"']/g,ch=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[ch]))}
function iso(d){let z=n=>String(n).padStart(2,"0");return d.getFullYear()+"-"+z(d.getMonth()+1)+"-"+z(d.getDate())}
function jsonSetting(name,fallback){try{return JSON.parse(settings()[name]||"")}catch{return fallback}}
function todayIso(){return iso(new Date())}
function paymentLabel(v){return ({cash:"Cash",qr:"QR",card:"Card",online:"Online Booking",transfer:"Bank Transfer"}[v]||v)}
const ROOM_GROUPS={suite:["101","201","301"],twin:["105-102","112-109","205-202","212-209","305-302","312-309"],family:["106-108","113-115","206-208","213-215","306-308","313-315"]};
function expandRooms(list){let out=[];list.forEach(item=>{if(!String(item).includes("-")){out.push(String(item));return}let [a,b]=String(item).split("-").map(Number),step=a<=b?1:-1;for(let n=a;;n+=step){out.push(String(n));if(n===b)break}});return out}
const ROOM_LOOKUP=Object.fromEntries(Object.entries(ROOM_GROUPS).flatMap(([type,list])=>expandRooms(list).map(room=>[room,type])));
function roomCategory(room){return ROOM_LOOKUP[String(room||"").trim()]||"twin"}
function manualBaseRate(){let s=settings(),stayType=q("#manualStayType").value,room=q("#manualRoom").value.trim(),category=roomCategory(room);if(stayType==="3_hour")return asMoney(s.rate_day_use||"65");if(category==="suite")return asMoney(s.rate_suite||"168");if(category==="family")return asMoney(s.rate_family||"148");return asMoney(s.rate_twin||"108")}
function manualComputedAmount(){let registration=q("#manualRegistration").value,duration=Math.max(1,Number(q("#manualDuration").value||1));if(registration==="online")return asMoney(q("#manualAmount").value);let base=manualBaseRate();return q("#manualStayType").value==="3_hour"?base:base*duration}
function syncManualDates(){let start=q("#manualCheckIn").value||todayIso(),stayType=q("#manualStayType").value,duration=Math.max(1,Number(q("#manualDuration").value||1));q("#manualCheckIn").value=start;if(stayType==="3_hour"){q("#manualDuration").value=1;q("#manualDuration").disabled=true}else{q("#manualDuration").disabled=false}let d=new Date(start+"T00:00:00");d.setDate(d.getDate()+(stayType==="3_hour"?0:duration));q("#manualCheckOut").value=iso(d)}
function setManualPaymentDefaults(){let payment=q("#manualPayment").value;q("#manualKelestarianPayment").value=payment;q("#manualDepositPayment").value=payment}
function syncManualPayments(){let registration=q("#manualRegistration").value;if(registration!=="online"&&!q("#manualAmount").dataset.manual)q("#manualAmount").value=manualComputedAmount().toFixed(2);if(registration==="online"&&!q("#manualAmount").value)q("#manualAmount").placeholder="Manual price";let nights=q("#manualStayType").value==="3_hour"?1:Math.max(1,Number(q("#manualDuration").value||1)),kel=asMoney(settings().fee_rate||"5.00")*nights,deposit=asMoney(q("#manualDeposit").value||"50"),room=q("#manualRoom").value.trim(),category=roomCategory(room).replace("suite","family suite");q("#manualPreview").className="";q("#manualPreview").innerHTML=`${registration==="online"?"Online booking manual price":category} · Room charge RM ${asMoney(q("#manualAmount").value).toFixed(2)} · Kelestarian RM ${kel.toFixed(2)} (${paymentLabel(q("#manualKelestarianPayment").value)}) · Deposit RM ${deposit.toFixed(2)} (${paymentLabel(q("#manualDepositPayment").value)})`}
function syncManualForm(){syncManualDates();syncManualPayments()}
function rangeBounds(){let now=new Date(),start=new Date(now),end=new Date(now);start.setHours(0,0,0,0);end.setHours(0,0,0,0);if(rangeMode==="last7"){start.setDate(start.getDate()-6)}else if(rangeMode==="month"){start=new Date(now.getFullYear(),now.getMonth(),1);end=new Date(now.getFullYear(),now.getMonth()+1,0)}else if(rangeMode==="custom"){let s=q("#startDate").value,e=q("#endDate").value;return {start:s||"0000-01-01",end:e||"9999-12-31"}}return {start:iso(start),end:iso(end)}}
function dailySalesBounds(){if(dailySalesMode==="custom"){let s=q("#dailyStartDate").value,e=q("#dailyEndDate").value;return {start:s||todayIso(),end:e||s||todayIso()}}let t=todayIso();return {start:t,end:t}}
function dataBounds(){let dates=sales.map(r=>r.date).filter(Boolean).sort();return dates.length?{start:dates[0],end:dates.at(-1)}:null}
function showImportedRange(bounds){bounds=bounds||dataBounds();if(!bounds)return "";rangeMode="custom";q("#startDate").value=bounds.start;q("#endDate").value=bounds.end;localStorage.setItem("rangeMode",rangeMode);localStorage.setItem("startDate",bounds.start);localStorage.setItem("endDate",bounds.end);return ` Showing ${bounds.start.split("-").reverse().join("/")} to ${bounds.end.split("-").reverse().join("/")}.`}
function visibleRows(){let bounds=rangeBounds(),term=q("#search").value.toLowerCase();return sales.filter(r=>r.date>=bounds.start&&r.date<=bounds.end).filter(r=>(String(r.guest_name||"")+" "+String(r.room_no||"")).toLowerCase().includes(term))}
function manualTransactionRows(bounds){bounds=bounds||dailySalesBounds();return sales.filter(r=>!r.is_paid_continuation&&(r.flags||[]).includes("MANUAL_CHECK_IN")).filter(r=>(r.check_in_date||r.date)>=bounds.start&&(r.check_in_date||r.date)<=bounds.end).sort((a,b)=>String(b.input_time||"").localeCompare(String(a.input_time||"")))}
function groupRows(rows){let html='<div class="table-wrap"><table><thead><tr><th>Room</th><th>Guest</th><th>Stay</th><th>Room Payment</th><th>Lestari Fee</th><th>Lestari Payment Method</th><th>Deposit</th><th>Deposit Payment Method</th><th>Room Payment Method</th><th>Rate Type</th></tr></thead><tbody>';let cur='';rows.forEach(r=>{let paid=r.is_paid_continuation?`<span class="paid-pill">${r.payment_status||("Paid on "+(r.display_date||""))}</span>`:`RM ${r.price||"0.00"}`;let fee=r.is_paid_continuation?"":`RM ${r.kelestarian||"0.00"}`;let deposit=r.deposit?`RM ${r.deposit}`:"";if(r.display_date!==cur){cur=r.display_date;html+=`<tr class="day-row"><td colspan="10">${cur}</td></tr>`}html+=`<tr class="${r.is_paid_continuation?'paid-continuation':(r.multi_night?'multi':'')}"><td>${r.room_no||''}</td><td>${r.guest_name||''}</td><td>${r.stay_progress||''}</td><td>${paid}</td><td>${fee}</td><td>${r.kelestarian_payment_method||''}</td><td>${deposit}</td><td>${r.deposit_payment_method||''}</td><td>${r.payment_method||''}</td><td>${r.rate_type||''}</td></tr>`});return html+'</tbody></table></div>'}
function flagValue(flags,key){let prefix=key+":";let found=(flags||[]).find(f=>String(f).startsWith(prefix));return found?String(found).slice(prefix.length):""}
function formatDateTime(value){if(!value)return "";let d=new Date(value);return Number.isNaN(d.getTime())?String(value).slice(0,19).replace("T"," "):d.toLocaleString("en-MY",{year:"numeric",month:"2-digit",day:"2-digit",hour:"2-digit",minute:"2-digit"})}
function renderManualHistory(){let selected=q("#manualHistoryDate")?q("#manualHistoryDate").value:"",rows=sales.filter(r=>!r.is_paid_continuation&&(r.flags||[]).includes("MANUAL_CHECK_IN")).filter(r=>!selected||(r.check_in_date||r.date)===selected).sort((a,b)=>String(b.input_time||"").localeCompare(String(a.input_time||"")));let el=q("#manualHistory");if(!el)return;el.className=rows.length?'':'empty';el.innerHTML=rows.length?table(rows,['Input time','Room','Registration','Stay type','Duration','Check in','Check out','Room payment','Kelestarian payment','Deposit payment','Remark',''],r=>`<tr><td>${html(formatDateTime(r.input_time||flagValue(r.flags,"INPUT_TIME")))}</td><td>${html(r.room_no)}</td><td>${html(r.registration||flagValue(r.flags,"REGISTRATION"))}</td><td>${html(r.stay_type_label||flagValue(r.flags,"STAY_TYPE"))}</td><td>${html(r.stay_progress)}</td><td>${html(String(r.check_in_date||r.date).split("-").reverse().join("/"))}</td><td>${html(String(r.check_out_date||"").split("-").reverse().join("/"))}</td><td>RM ${r.price||"0.00"} (${html(r.payment_method||"")})</td><td>RM ${r.kelestarian||"0.00"} (${html(r.kelestarian_payment_method||"")})</td><td>RM ${r.deposit||"0.00"} (${html(r.deposit_payment_method||"")})</td><td>${html(r.remark||flagValue(r.flags,"REMARK"))}</td><td><button class="danger" data-delete-manual="${html(r.folio_no||"")}">Delete</button></td></tr>`):selected?`No manual check-ins saved for ${html(selected.split("-").reverse().join("/"))}.`:'No manual check-ins saved yet.';qa("[data-delete-manual]").forEach(btn=>btn.onclick=()=>deleteManualTransaction(btn.dataset.deleteManual))}
function cashDepositActive(row){if((row.deposit_payment_method||"").toLowerCase()!=="cash")return false;let rowDate=row.check_in_date||row.date,today=todayIso();if(rowDate!==today)return true;let input=new Date((row.input_time||flagValue(row.flags,"INPUT_TIME")||"")+"");if(Number.isNaN(input.getTime()))return true;let reset=new Date();reset.setHours(12,0,0,0);return new Date() < reset || input >= reset}
function methodKey(value){let text=String(value||"").toLowerCase();if(text.includes("cash"))return "cash";if(text.includes("qr"))return "qr";if(text.includes("transfer")||text.includes("bank"))return "transfer";if(text.includes("card"))return "card";if(text.includes("online"))return "online";return ""}
function moneyCard(label,value,extraClass=""){return `<article class="metric ${extraClass}"><span>${label}</span><strong>RM ${value.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</strong></article>`}
function cashflowBucket(value){let key=methodKey(value);if(key==="cash")return "cash";if(key==="qr")return "qr";if(["transfer","card","online"].includes(key))return "bank";return ""}
function dailyCheckInRows(bounds){bounds=bounds||dailySalesBounds();return sales.filter(r=>!r.is_paid_continuation).filter(r=>(r.check_in_date||r.date)>=bounds.start&&(r.check_in_date||r.date)<=bounds.end).sort((a,b)=>String(b.input_time||b.date||"").localeCompare(String(a.input_time||a.date||"")))}
function ownerRows(bounds){bounds=bounds||dailySalesBounds();return ownerMovements.filter(r=>r.date>=bounds.start&&r.date<=bounds.end).sort((a,b)=>String(b.created_at||"").localeCompare(String(a.created_at||"")))}
function transferRows(bounds){bounds=bounds||dailySalesBounds();return transferMovements.filter(r=>r.date>=bounds.start&&r.date<=bounds.end).sort((a,b)=>String(b.created_at||"").localeCompare(String(a.created_at||"")))}
function emptyCashflow(){return {room:0,lestari:0,deposit:0,equity:0,drawing:0,transfer_in:0,transfer_out:0}}
function renderDailyCashflow(rows,bounds){let totals={cash:emptyCashflow(),qr:emptyCashflow(),bank:emptyCashflow()};rows.forEach(r=>{let roomBucket=cashflowBucket(r.payment_method),lestariBucket=cashflowBucket(r.kelestarian_payment_method||r.payment_method),depositBucket=cashflowBucket(r.deposit_payment_method);if(roomBucket)totals[roomBucket].room+=asMoney(r.price);if(lestariBucket)totals[lestariBucket].lestari+=asMoney(r.kelestarian);if(depositBucket&&!(depositBucket==="cash"&&!cashDepositActive(r)))totals[depositBucket].deposit+=asMoney(r.deposit)});ownerRows(bounds).forEach(r=>{let bucket=cashflowBucket(r.method);if(!bucket)return;totals[bucket][r.type==="drawing"?"drawing":"equity"]+=asMoney(r.amount)});transferRows(bounds).forEach(r=>{let from=cashflowBucket(r.from),to=cashflowBucket(r.to),amount=asMoney(r.amount);if(from)totals[from].transfer_out+=amount;if(to)totals[to].transfer_in+=amount});let section=(title,key,includeDeposit)=>{let t=totals[key],final=t.room+t.lestari+(includeDeposit?t.deposit:0)+t.equity-t.drawing+t.transfer_in-t.transfer_out,finalLabel=title==="Bank"?"Bank value":"Final on hand";return `<h2>${title}</h2><br><div class="cards">${moneyCard("Room payment",t.room)}${moneyCard("Lestari",t.lestari)}${includeDeposit?moneyCard("Deposit",t.deposit):""}</div><div class="final-row">${moneyCard(finalLabel,final,"final")}</div>`};let el=q("#dailyCashflow");if(!el)return;el.className=(rows.length||ownerRows(bounds).length||transferRows(bounds).length)?'':'empty';el.innerHTML=(rows.length||ownerRows(bounds).length||transferRows(bounds).length)?section("Cash","cash",true)+section("QR","qr",true)+section("Bank","bank",false):'No check-ins in this date range.'}
function selectedOwnerName(){let selected=q("#ownerSelect").value;return selected==="__new__"?q("#ownerName").value.trim():selected}
function saveOwnerMovement(){let name=selectedOwnerName(),amount=asMoney(q("#ownerAmount").value),bounds=dailySalesBounds();if(!name||amount<=0){q("#notice").textContent="Choose or add an owner, then add an amount first.";return}let date=bounds.start===bounds.end?bounds.start:todayIso();ownerMovements.push({id:String(Date.now()),date,created_at:new Date().toISOString(),name,type:q("#ownerType").value,method:q("#ownerMethod").value,amount:amount.toFixed(2),note:q("#ownerNote").value.trim()});localStorage.setItem("ownerMovements",JSON.stringify(ownerMovements));q("#ownerAmount").value="";q("#ownerNote").value="";q("#ownerFilter").value=name;q("#ownerName").value="";q("#notice").textContent=`${q("#ownerType").value==="drawing"?"Drawing":"Owner equity"} added for ${name}.`;render()}
function deleteOwnerMovement(id){ownerMovements=ownerMovements.filter(r=>r.id!==id);localStorage.setItem("ownerMovements",JSON.stringify(ownerMovements));render()}
function ownerNames(){return [...new Set(ownerMovements.map(r=>String(r.name||"").trim()).filter(Boolean))].sort((a,b)=>a.localeCompare(b))}
function syncOwnerPicker(){let names=ownerNames(),filter=q("#ownerFilter"),select=q("#ownerSelect"),wrap=q("#ownerNameWrap"),filterSelected=filter?filter.value:"",selectSelected=select?select.value:"";let ownerOptions=names.map(n=>`<option value="${html(n)}">${html(n)}</option>`).join("");if(filter){filter.innerHTML='<option value="">All owners</option>'+ownerOptions;filter.value=names.includes(filterSelected)?filterSelected:""}if(select){select.innerHTML='<option value="">Choose owner</option>'+ownerOptions+'<option value="__new__">+ Add new owner</option>';select.value=names.includes(selectSelected)||selectSelected==="__new__"?selectSelected:""}if(wrap)wrap.style.display=select&&select.value==="__new__"?"grid":"none"}
function renderOwnerMovements(bounds){syncOwnerPicker();let selected=q("#ownerFilter")?q("#ownerFilter").value:"",rows=ownerMovements.filter(r=>!selected||String(r.name||"")===selected).sort((a,b)=>String(b.date||"").localeCompare(String(a.date||""))||String(b.created_at||"").localeCompare(String(a.created_at||""))),el=q("#ownerMovements");if(!el)return;el.className=rows.length?'':'empty';el.innerHTML=rows.length?table(rows,['Date','Owner','Type','Method','Amount','Remark',''],r=>`<tr><td>${html(String(r.date||"").split("-").reverse().join("/"))}</td><td>${html(r.name)}</td><td>${html(r.type==="drawing"?"Drawings":"Owner Equity")}</td><td>${html(paymentLabel(r.method))}</td><td>RM ${asMoney(r.amount).toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td><td>${html(r.note||"")}</td><td><button class="danger" data-delete-owner="${html(r.id)}">Delete</button></td></tr>`):selected?`No owner equity or drawings history for ${html(selected)}.`:'No owner equity or drawings history yet.';qa("[data-delete-owner]").forEach(btn=>btn.onclick=()=>deleteOwnerMovement(btn.dataset.deleteOwner))}
function saveTransferMovement(){let from=q("#transferFrom").value,to=q("#transferTo").value,amount=asMoney(q("#transferAmount").value),bounds=dailySalesBounds();if(from===to){q("#notice").textContent="Transfer source and destination must be different.";return}if(amount<=0){q("#notice").textContent="Add transfer amount first.";return}let date=bounds.start===bounds.end?bounds.start:todayIso();transferMovements.push({id:String(Date.now()),date,created_at:new Date().toISOString(),from,to,amount:amount.toFixed(2),note:q("#transferNote").value.trim()});localStorage.setItem("transferMovements",JSON.stringify(transferMovements));q("#transferAmount").value="";q("#transferNote").value="";q("#notice").textContent="Transfer added.";render()}
function deleteTransferMovement(id){transferMovements=transferMovements.filter(r=>r.id!==id);localStorage.setItem("transferMovements",JSON.stringify(transferMovements));render()}
function renderTransferMovements(bounds){let rows=transferRows(bounds),el=q("#transferMovements");if(!el)return;el.className=rows.length?'':'empty';el.innerHTML=rows.length?table(rows,['Date','From','To','Amount','Remark',''],r=>`<tr><td>${html(String(r.date||"").split("-").reverse().join("/"))}</td><td>${html(paymentLabel(r.from))}</td><td>${html(paymentLabel(r.to))}</td><td>RM ${asMoney(r.amount).toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td><td>${html(r.note||"")}</td><td><button class="danger" data-delete-transfer="${html(r.id)}">Delete</button></td></tr>`):'No transfers for this date range.';qa("[data-delete-transfer]").forEach(btn=>btn.onclick=()=>deleteTransferMovement(btn.dataset.deleteTransfer))}
function renderDailySales(){let bounds=dailySalesBounds(),rows=dailyCheckInRows(bounds);renderDailyCashflow(rows,bounds);renderOwnerMovements(bounds);renderTransferMovements(bounds);let el=q("#dailySalesTable");if(!el)return;el.className=rows.length?'':'empty';el.innerHTML=rows.length?table(rows,['Input time','Room','Guest','Registration','Stay','Check in','Check out','Room payment','Kelestarian payment','Deposit payment','Payment','Folio','Bill','Remark'],r=>`<tr><td>${html(formatDateTime(r.input_time||flagValue(r.flags,"INPUT_TIME")))}</td><td>${html(r.room_no)}</td><td>${html(r.guest_name)}</td><td>${html(r.registration||flagValue(r.flags,"REGISTRATION")||r.rate_type||"")}</td><td>${html(r.stay_progress)}</td><td>${html(String(r.check_in_date||r.date).split("-").reverse().join("/"))}</td><td>${html(String(r.check_out_date||"").split("-").reverse().join("/"))}</td><td>RM ${r.price||"0.00"} (${html(r.payment_method||"")})</td><td>RM ${r.kelestarian||"0.00"} (${html(r.kelestarian_payment_method||r.payment_method||"")})</td><td>${r.deposit?`RM ${r.deposit} (${html(r.deposit_payment_method||"")})`:""}</td><td>${html(r.payment_method||"")}</td><td>${html(r.folio_no||"")}</td><td>${html(r.bill_no||"")}</td><td>${html(r.remark||flagValue(r.flags,"REMARK"))}</td></tr>`):'No check-ins in this date range.'}
function table(rows,heads,mapper){return '<div class="table-wrap"><table><thead><tr>'+heads.map(h=>`<th>${h}</th>`).join('')+'</tr></thead><tbody>'+rows.map(mapper).join('')+'</tbody></table></div>'}
function reportCards(rows,mapper){return '<div class="report-card-grid">'+rows.map(mapper).join('')+'</div>'}
function reportLine(label,value){return `<div><dt>${label}</dt><dd>${value}</dd></div>`}
function collectionBreakdown(rows){let map={};rows.filter(r=>!r.is_paid_continuation).forEach(r=>{let key=r.payment_method||"Unknown";map[key]??={payment_method:key,bills:0,amount:0,kelestarian:0};map[key].bills++;map[key].amount+=asMoney(r.price);map[key].kelestarian+=asMoney(r.kelestarian)});return Object.values(map).sort((a,b)=>b.amount-a.amount)}
function renderHistorical(data){let el=q("#historicalPanel");if(!data||!data.summary){el.className='empty';el.innerHTML='No historical data loaded.';return}let h=data.summary,batches=data.archive_batches||[],picker=q("#historicalYear"),years=h.available_years||[],selected=h.selected_year||(years.at(-1)||'');if(picker&&years.length){picker.innerHTML=years.map(y=>`<option value="${y}">${y}</option>`).join('');picker.value=selected;localStorage.setItem("historicalYear",selected)}el.className='';el.innerHTML='<div class="cards"><article class="metric"><span>Historical stays '+selected+'</span><strong>'+Number(h.historical_stays||0).toLocaleString()+'</strong></article><article class="metric"><span>Historical revenue '+selected+'</span><strong>RM '+(h.total_revenue||'0.00')+'</strong></article><article class="metric"><span>Room nights '+selected+'</span><strong>'+Number(h.total_nights||0).toLocaleString()+'</strong></article><article class="metric"><span>2026 estimate</span><strong>RM '+(h.forecast_2026_kelestarian||'0.00')+'</strong></article></div><h2>'+selected+' sales</h2><br>'+table(h.yearly_revenue||[],['Year','Stays','Nights','Revenue'],r=>`<tr><td>${r.year}</td><td>${Number(r.stays).toLocaleString()}</td><td>${Number(r.nights).toLocaleString()}</td><td>RM ${r.revenue}</td></tr>`)+'<br><h2>Monthly trend '+selected+'</h2><br>'+table(h.monthly_revenue||[],['Month','Stays','Nights','Revenue'],r=>`<tr><td>${r.month}</td><td>${Number(r.stays).toLocaleString()}</td><td>${Number(r.nights).toLocaleString()}</td><td>RM ${r.revenue}</td></tr>`)+'<br><h2>Payment categories '+selected+'</h2><br>'+table(h.payment_categories||[],['Payment','Bills','Amount'],r=>`<tr><td>${r.payment_method}</td><td>${Number(r.count).toLocaleString()}</td><td>RM ${r.total_amount}</td></tr>`)+'<br><h2>Guest memory '+selected+'</h2><br>'+table(h.top_guests||[],['Guest','Stays','Nights','Spend','First','Latest'],r=>`<tr><td>${html(r.guest_name)}</td><td>${r.stays}</td><td>${r.nights}</td><td>RM ${r.total_amount}</td><td>${r.first}</td><td>${r.latest}</td></tr>`)+'<br><h2>Data quality issues '+selected+'</h2><br>'+((h.data_quality_issues||[]).length?table(h.data_quality_issues||[],['Bill','Guest','Room','Issue'],r=>`<tr><td>${html(r.bill_no)}</td><td>${html(r.guest_name)}</td><td>${html(r.room_no)}</td><td>${html((r.flags||[]).join(', '))}</td></tr>`):'<div class="empty">No historical data quality issues found for '+selected+'.</div>')+'<br><h2>Import archive</h2><br>'+table(batches.slice(0,80),['File','Imported','Stays','Revenue','Kelestarian'],r=>`<tr><td>${html(r.source_filename)}</td><td>${html(String(r.imported_at||'').slice(0,19).replace('T',' '))}</td><td>${Number(r.stay_count||0).toLocaleString()}</td><td>RM ${asMoney(r.total_sales).toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td><td>RM ${asMoney(r.total_kelestarian).toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td></tr>`)}
function monthEnd(month){let [y,m]=month.split("-").map(Number),d=new Date(y,m,0),z=n=>String(n).padStart(2,"0");return `${y}-${z(m)}-${z(d.getDate())}`}
function syncReportMonthC(){let month=q("#reportMonthC").value;if(!month)return;q("#reportStartC").value=`${month}-01`;q("#reportEndC").value=monthEnd(month);localStorage.setItem("reportStartC",q("#reportStartC").value);localStorage.setItem("reportEndC",q("#reportEndC").value)}
function setReportDefaults(){let dates=sales.filter(r=>r.date&&r.date>='2026-01-01'&&r.date<='2026-12-31'&&!r.is_paid_continuation).map(r=>r.date).sort(),latest=dates.at(-1),latestMonth=latest?latest.slice(0,7):todayIso().slice(0,7),savedB=localStorage.getItem("reportMonthB"),savedC=localStorage.getItem("reportMonthC"),cMode=localStorage.getItem("reportCMode")||"month";if(!q("#reportMonthB").value)q("#reportMonthB").value=savedB||latestMonth;if(cMode==="custom"){if(!q("#reportStartC").value)q("#reportStartC").value=localStorage.getItem("reportStartC")||`${latestMonth}-01`;if(!q("#reportEndC").value)q("#reportEndC").value=localStorage.getItem("reportEndC")||monthEnd(latestMonth);q("#reportMonthC").value="";return}if(!q("#reportMonthC").value)q("#reportMonthC").value=savedC||q("#reportMonthB").value||latestMonth;if(!q("#reportStartC").value||!q("#reportEndC").value||q("#reportStartC").value.slice(0,7)!==q("#reportMonthC").value||q("#reportEndC").value.slice(0,7)!==q("#reportMonthC").value)syncReportMonthC()}
function selectedCheckIns(kind){let rows=sales.filter(r=>!r.is_paid_continuation);if(kind==='b'){let month=q("#reportMonthB").value;return rows.filter(r=>r.date.slice(0,7)===month)}let month=q("#reportMonthC").value;if(month)return rows.filter(r=>r.date.slice(0,7)===month);let start=q("#reportStartC").value,end=q("#reportEndC").value;return rows.filter(r=>(!start||r.date>=start)&&(!end||r.date<=end))}
function renderReviewQueue(){let el=q("#reviewQueue");el.className=reviewQueue.length?'':'empty';el.innerHTML=reviewQueue.length?table(reviewQueue,['Folio No.','Incoming Bill','Existing Bill','Guest','Reason','Source'],r=>`<tr><td>${html(r.folio_no)}</td><td>${html(r.incoming_bill_no)}</td><td>${html(r.existing_bill_no)}</td><td>${html((r.incoming_record||{}).guest_name)}</td><td class="review-reason">${html(String(r.reason||'').replaceAll('_',' '))}</td><td>${html(r.source_filename)}</td></tr>`):'No records need manual review.'}
function render(){setReportDefaults();qa(".filters button").forEach(b=>b.classList.toggle("active",b.dataset.range===rangeMode));qa("[data-daily-range]").forEach(b=>b.classList.toggle("active",b.dataset.dailyRange===dailySalesMode));let filtered=visibleRows();let visibleSales=filtered.reduce((t,r)=>t+asMoney(r.price),0),visibleFee=filtered.reduce((t,r)=>t+asMoney(r.kelestarian),0);q("#mStays").textContent=summary.stays||0;q("#mRows").textContent=filtered.length;q("#mSales").textContent='RM '+visibleSales.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2});q("#mFee").textContent='RM '+visibleFee.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2});q("#downloadB").disabled=!stays.length;q("#downloadC").disabled=!stays.length;q("#ledger").className=filtered.length?'':'empty';q("#ledger").innerHTML=filtered.length?groupRows(filtered):'No rows in this date range.';renderManualHistory();renderDailySales();let bSelected=selectedCheckIns('b'),byDay={};bSelected.forEach(r=>{byDay[r.display_date]??={rooms:0,nights:0,fee:0};byDay[r.display_date].rooms++;byDay[r.display_date].nights+=Number(r.nights||0);byDay[r.display_date].fee+=asMoney(r.kelestarian)});let bRows=Object.entries(byDay).map(([d,v])=>({d,...v}));q("#previewB").className=bRows.length?'':'empty';q("#previewB").innerHTML=bRows.length?table(bRows,['Tarikh','Jumlah Bilik','Bilangan Malam','Jumlah Kutipan'],r=>`<tr><td>${r.d}</td><td>${r.rooms}</td><td>${r.nights}</td><td>RM ${r.fee.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td></tr>`):'No check-ins in the selected month.';let cSelected=selectedCheckIns('c');q("#previewC").className=cSelected.length?'':'empty';q("#previewC").innerHTML=cSelected.length?groupRows(cSelected):'No check-ins in the selected date range.';let analytics=q("#analytics"),collections=collectionBreakdown(filtered);if(!summary.payment_method_breakdown){analytics.className='empty';analytics.innerHTML='Import Excel first.'}else{analytics.className='';analytics.innerHTML='<div class="cards"><article class="metric"><span>Room payments in selected dates</span><strong>RM '+visibleSales.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})+'</strong></article><article class="metric"><span>Lestari fees in selected dates</span><strong>RM '+visibleFee.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})+'</strong></article><article class="metric"><span>Room payment methods used</span><strong>'+collections.length+'</strong></article><article class="metric"><span>Import data issues</span><strong>'+(summary.data_quality_issues||[]).length+'</strong></article></div><h2>Room payment collection by method</h2><br>'+ (collections.length?table(collections,['Room Payment Method','Transactions','Room Payment Collected','Lestari Fee'],r=>`<tr><td>${r.payment_method}</td><td>${r.bills}</td><td>RM ${r.amount.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td><td>RM ${r.kelestarian.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</td></tr>`):'<div class="empty">No collections in this date range.</div>') + '<br><h2>Revenue by rate type</h2><br>'+table(summary.revenue_by_rate_type,['Rate Type','Bills','Amount'],r=>`<tr><td>${r.rate_type}</td><td>${r.count}</td><td>RM ${r.total_amount}</td></tr>`) }}
async function importFile(file){q("#dropText").innerHTML=html(file.name)+'<small>Importing...</small>';let fd=new FormData();fd.append('file',file);fd.append('fee_rate',settings().fee_rate||'5.00');let res=await fetch('/api/import',{method:'POST',body:fd});let data=await res.json();if(!res.ok){q("#notice").textContent=data.error||'Import failed';q("#dropText").innerHTML=html(file.name)+'<small>Import failed</small>';return}stays=data.stays;sales=data.sales;summary=data.summary;localStorage.setItem('stays',JSON.stringify(stays));localStorage.setItem('sales',JSON.stringify(sales));localStorage.setItem('summary',JSON.stringify(summary));let rangeNote=showImportedRange(data.import_start&&data.import_end?{start:data.import_start,end:data.import_end}:null);let savedNote=data.saved?`${data.inserted_count||0} new, ${data.updated_count||0} updated`:`${data.accepted_count||0} parsed`;let message=`Imported ${summary.stays} stays. ${savedNote}.`+rangeNote;if(data.merged_duplicate_count)message+=` ${data.merged_duplicate_count} duplicate row(s) merged.`;if(data.review_count)message+=` ${data.review_count} item(s) logged for review.`;if(data.review_error)message+=` Manual Review warning: ${data.review_error}`;if(data.save_error)message+=` Database warning: ${data.save_error}`;q("#notice").textContent=message;q("#dropText").innerHTML=html(file.name)+'<small>Imported</small>';render();loadReviewQueue()}
async function deleteManualTransaction(folioNo){if(!folioNo){q("#notice").textContent="This manual transaction cannot be deleted because its reference is missing.";return}if(!confirm("Delete this manual transaction?"))return;let fd=new FormData();fd.append("folio_no",folioNo);let res=await fetch("/api/manual-checkin/delete",{method:"POST",body:fd}),data=await res.json();if(!res.ok){q("#notice").textContent=data.error||"Delete failed.";return}stays=stays.filter(s=>s.folio_no!==folioNo);sales=sales.filter(r=>r.folio_no!==folioNo);localStorage.setItem("stays",JSON.stringify(stays));localStorage.setItem("sales",JSON.stringify(sales));q("#notice").textContent="Manual transaction deleted.";render();await loadHistory(rangeBounds())}
async function saveManualCheckIn(){syncManualForm();let checkIn=q("#manualCheckIn").value||todayIso(),fd=new FormData();qa(".manual").forEach(i=>fd.append(i.name,i.value));fd.append("fee_rate",settings().fee_rate||"5.00");let res=await fetch('/api/manual-checkin',{method:'POST',body:fd}),data=await res.json();if(!res.ok){q("#notice").textContent=data.error||'Manual check-in failed.';return}q("#notice").textContent=`Manual check-in saved for room ${html(q("#manualRoom").value)}.`;q("#manualRoom").value="";q("#manualGuest").value="Walk-in guest";q("#manualRemark").value="";q("#manualAmount").dataset.manual="";q("#manualAmount").value="";q("#manualHistoryDate").value=checkIn;syncManualForm();await loadHistory({start:checkIn,end:checkIn});loadReviewQueue()}
async function loadHistory(bounds=null){let params=new URLSearchParams({fee_rate:settings().fee_rate||'5.00'});if(bounds){params.set('start',bounds.start);params.set('end',bounds.end)}let res=await fetch('/api/history?'+params.toString());let data=await res.json();if(!res.ok||!data.enabled){q("#notice").textContent=data.error||'Saved stays are not available.';return}stays=data.stays||[];sales=data.sales||[];summary=data.summary||{};localStorage.setItem('stays',JSON.stringify(stays));localStorage.setItem('sales',JSON.stringify(sales));localStorage.setItem('summary',JSON.stringify(summary));let rangeNote=visibleRows().length?"":showImportedRange();q("#notice").textContent=`Loaded ${summary.stays||0} saved stays from Supabase.`+rangeNote;render()}
async function loadHistorical(){let year=q("#historicalYear").value||localStorage.getItem("historicalYear")||"2025";q("#historicalPanel").className='empty';q("#historicalPanel").innerHTML='Loading '+year+' historical data...';let url='/api/historical?fee_rate='+(settings().fee_rate||'5.00')+'&year='+encodeURIComponent(year);let res=await fetch(url);let data=await res.json();if(!res.ok||!data.enabled){q("#historicalPanel").innerHTML=data.error||'Historical data is not available.';return}renderHistorical(data)}
async function loadReviewQueue(){let res=await fetch('/api/review-queue'),data=await res.json();reviewQueue=res.ok&&data.enabled?(data.records||[]):[];renderReviewQueue()}
async function download(kind){let fd=new FormData();fd.append('kind',kind);fd.append('stays',JSON.stringify(stays));if(kind==='b'){let month=q("#reportMonthB").value;if(!month){q("#notice").textContent='Select a month for Lampiran B.';return}fd.append('report_month',month)}else{let month=q("#reportMonthC").value;if(month){syncReportMonthC();fd.append('report_month',month)}let start=q("#reportStartC").value,end=q("#reportEndC").value;if(!month&&(!start||!end)){q("#notice").textContent='Select a month or date range for Lampiran C.';return}if(start&&end&&start>end){q("#notice").textContent='Lampiran C first date must not be after the last date.';return}if(!month){fd.append('report_start',start);fd.append('report_end',end)}}Object.entries(settings()).forEach(([k,v])=>fd.append(k,v));let res=await fetch('/api/report',{method:'POST',body:fd});if(!res.ok){q("#notice").textContent=await res.text();return}let blob=await res.blob(),url=URL.createObjectURL(blob),a=document.createElement('a'),disposition=res.headers.get('Content-Disposition')||'',match=disposition.match(/filename="?([^";]+)"?/i);a.href=url;a.download=match?match[1]:(kind==='b'?'Lampiran B.pdf':'Lampiran C.pdf');a.click();URL.revokeObjectURL(url)}
const WORKSPACE_DEFAULT_VIEW={sales:"importSales",dashboard:"dashboard",lestari:"b"};
function viewWorkspace(view){let btn=q(`.nav button[data-view="${view}"]`);return btn?btn.dataset.workspace:"sales"}
function setWorkspace(viewOrWorkspace){let workspace=WORKSPACE_DEFAULT_VIEW[viewOrWorkspace]?viewOrWorkspace:viewWorkspace(viewOrWorkspace),view=WORKSPACE_DEFAULT_VIEW[viewOrWorkspace]||viewOrWorkspace,button=q(`.nav button[data-view="${view}"]`)||q(`.nav button[data-workspace="${workspace}"]`);if(!button)return;workspace=button.dataset.workspace;view=button.dataset.view;qa(".nav button").forEach(x=>{let visible=x.dataset.workspace===workspace;x.style.display=visible?"block":"none";x.classList.toggle('active',visible&&x.dataset.view===view)});qa(".view").forEach(v=>v.classList.remove('active'));q("#"+view).classList.add('active');q("#pageTitle").textContent=button.textContent;q("#workspaceSelect").value=workspace}
qa(".nav button").forEach(b=>b.onclick=()=>setWorkspace(b.dataset.view));
q("#workspaceSelect").onchange=e=>setWorkspace(e.target.value);
q("#file").onchange=e=>e.target.files[0]&&importFile(e.target.files[0]);q("#search").oninput=render;q("#downloadB").onclick=()=>download('b');q("#downloadC").onclick=()=>download('c');
q("#saveManual").onclick=saveManualCheckIn;
q("#saveOwnerMovement").onclick=saveOwnerMovement;
q("#ownerFilter").onchange=()=>renderOwnerMovements(dailySalesBounds());
q("#ownerSelect").onchange=syncOwnerPicker;
q("#saveTransferMovement").onclick=saveTransferMovement;
q("#refreshReview").onclick=loadReviewQueue;
q("#loadHistorical").onclick=loadHistorical;
q("#historicalYear").onchange=loadHistorical;
qa(".filters button").forEach(b=>b.onclick=()=>{rangeMode=b.dataset.range;localStorage.setItem("rangeMode",rangeMode);render()});
q("#startDate").value=localStorage.getItem("startDate")||"";q("#endDate").value=localStorage.getItem("endDate")||"";
["#startDate","#endDate"].forEach(id=>q(id).onchange=()=>{rangeMode="custom";localStorage.setItem("rangeMode",rangeMode);localStorage.setItem(id.slice(1),q(id).value);loadHistory(rangeBounds())});
q("#dailyStartDate").value=localStorage.getItem("dailyStartDate")||todayIso();q("#dailyEndDate").value=localStorage.getItem("dailyEndDate")||todayIso();
qa("[data-daily-range]").forEach(b=>b.onclick=()=>{dailySalesMode=b.dataset.dailyRange;localStorage.setItem("dailySalesMode",dailySalesMode);if(dailySalesMode==="today"){q("#dailyStartDate").value=todayIso();q("#dailyEndDate").value=todayIso()}render()});
["#dailyStartDate","#dailyEndDate"].forEach(id=>q(id).onchange=()=>{dailySalesMode="custom";localStorage.setItem("dailySalesMode",dailySalesMode);localStorage.setItem(id.slice(1),q(id).value);render()});
q("#manualHistoryDate").onchange=()=>{let d=q("#manualHistoryDate").value;if(d)loadHistory({start:d,end:d});else renderManualHistory()};
q("#reportMonthB").onchange=()=>{localStorage.setItem("reportMonthB",q("#reportMonthB").value);render()};
q("#reportMonthB").value=localStorage.getItem("reportMonthB")||"";
q("#reportMonthC").value=localStorage.getItem("reportMonthC")||"";
q("#reportStartC").value=localStorage.getItem("reportStartC")||"";
q("#reportEndC").value=localStorage.getItem("reportEndC")||"";
["#reportStartC","#reportEndC"].forEach(id=>q(id).onchange=()=>{q("#reportMonthC").value="";localStorage.removeItem("reportMonthC");localStorage.setItem("reportCMode","custom");localStorage.setItem(id.slice(1),q(id).value);render()});
q("#reportMonthC").onchange=()=>{localStorage.setItem("reportCMode","month");localStorage.setItem("reportMonthC",q("#reportMonthC").value);syncReportMonthC();render()};
["#manualRoom","#manualRegistration","#manualStayType","#manualDuration","#manualCheckIn","#manualKelestarianPayment","#manualDepositPayment","#manualDeposit"].forEach(id=>q(id).oninput=syncManualForm);
q("#manualPayment").oninput=()=>{setManualPaymentDefaults();syncManualForm()};
q("#manualAmount").oninput=()=>{q("#manualAmount").dataset.manual="1";syncManualPayments()};
let drop=q("#drop");['dragenter','dragover'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.add('drag')}));['dragleave','drop'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.remove('drag')}));drop.addEventListener('drop',e=>{if(e.dataTransfer.files[0])importFile(e.dataTransfer.files[0])});
qa(".setting").forEach(i=>{i.value=localStorage.getItem('set_'+i.name)||i.value;i.oninput=()=>localStorage.setItem('set_'+i.name,i.value)});
q("#manualCheckIn").value=todayIso();q("#manualHistoryDate").value=todayIso();syncManualForm();
setWorkspace("sales");
render();
if(!stays.length)loadHistory();
loadReviewQueue();
</script>
</body>
</html>"""


@app.get("/")
def index():
    return PAGE


@app.post("/api/import")
def api_import():
    upload = request.files.get("file")
    if not upload:
        return jsonify({"error": "Please upload an Excel file."}), 400
    try:
        fee_rate = dec(request.form.get("fee_rate", "5.00"))
        stays = parse_stays(upload.filename, upload.read())
        sales = expanded_sales(stays, fee_rate)
        summary_data = summary(stays, fee_rate)
        dashboard_stays = kelestarian_stays(stays)
        dashboard_sales = ledger_sales(dashboard_stays, fee_rate)
        dashboard_summary = summary(dashboard_stays, fee_rate)
        save_result = {"saved": False, "accepted_count": 0, "review_count": 0}
        save_error = ""
        try:
            save_result = save_import_to_supabase(upload.filename, stays, sales, summary_data)
        except Exception as exc:
            save_error = str(exc)
        if supabase_configured() and save_error:
            return jsonify({"error": f"Database update failed. Import was not saved: {save_error}"}), 500
        import_dates = sorted({stay.check_out_date.isoformat() for stay in stays if stay.check_out_date})
        return jsonify({
            "stays": [stay_record(s) for s in dashboard_stays],
            "sales": dashboard_sales,
            "summary": dashboard_summary,
            "parsed_count": len(stays),
            "dashboard_count": len(dashboard_stays),
            "import_start": import_dates[0] if import_dates else "",
            "import_end": import_dates[-1] if import_dates else "",
            **save_result,
            "save_error": save_error,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/manual-checkin")
def api_manual_checkin():
    try:
        fee_rate = dec(request.form.get("fee_rate", "5.00"))
        stay = manual_stay_from_form(request.form)
        stays = [stay]
        sales = expanded_sales(stays, fee_rate)
        summary_data = summary(stays, fee_rate)
        save_result = {"saved": False, "accepted_count": 1, "review_count": 0}
        save_error = ""
        try:
            save_result = save_import_to_supabase("Manual daily check-in", stays, sales, summary_data)
        except Exception as exc:
            save_error = str(exc)
        if supabase_configured() and save_error:
            return jsonify({"error": f"Manual check-in was not saved: {save_error}"}), 500
        return jsonify({
            "stays": [stay_record(stay)],
            "sales": ledger_sales(stays, fee_rate),
            "summary": summary_data,
            **save_result,
            "save_error": save_error,
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/manual-checkin/delete")
def api_delete_manual_checkin():
    try:
        folio_no = verification_value(request.form.get("folio_no"))
        if not folio_no or not folio_no.startswith("MANUAL-"):
            return jsonify({"error": "Only manual transactions can be deleted here."}), 400
        if not supabase_configured():
            return jsonify({"enabled": False, "deleted": False, "folio_no": folio_no})
        supabase_request(
            "DELETE",
            "guest_stays",
            query={"folio_no": f"eq.{folio_no}"},
            prefer="return=minimal",
        )
        return jsonify({"enabled": True, "deleted": True, "folio_no": folio_no})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/history")
def api_history():
    try:
        if not supabase_configured():
            return jsonify({"enabled": False, "stays": [], "sales": [], "summary": {}})
        fee_rate = dec(request.args.get("fee_rate", "5.00"))
        try:
            start = date.fromisoformat(request.args.get("start", "") or KELESTARIAN_START.isoformat())
            end = date.fromisoformat(request.args.get("end", "") or KELESTARIAN_END.isoformat())
        except ValueError as exc:
            raise ValueError("Please select valid dashboard dates.") from exc
        if start > end:
            raise ValueError("Dashboard start date must be before or equal to the end date.")
        records = load_stays_by_checkin_range_from_supabase(start, end)
        if not records:
            return jsonify({"enabled": True, "stays": [], "sales": [], "summary": {}})
        stays = records_to_stays(records)
        return jsonify({"enabled": True, "stays": [stay_record(s) for s in stays], "sales": ledger_sales(stays, fee_rate), "summary": summary(stays, fee_rate)})
    except Exception as exc:
        return jsonify({"enabled": supabase_configured(), "error": str(exc)}), 400


@app.get("/api/historical")
def api_historical():
    try:
        if not supabase_configured():
            return jsonify({"enabled": False, "summary": {}, "archive_batches": []})
        fee_rate = dec(request.args.get("fee_rate", "5.00"))
        year_text = request.args.get("year", "")
        available_years = [str(year) for year in range(2016, KELESTARIAN_START.year)]
        selected_year = int(year_text) if year_text in available_years else KELESTARIAN_START.year - 1
        records = load_stays_by_checkout_range_from_supabase(date(selected_year, 1, 1), date(selected_year, 12, 31))
        stays = records_to_stays(records) if records else []
        summary_data = historical_summary(stays, fee_rate, selected_year)
        summary_data["available_years"] = available_years
        return jsonify({
            "enabled": True,
            "summary": summary_data,
            "archive_batches": load_import_batches_from_supabase(),
        })
    except Exception as exc:
        return jsonify({"enabled": supabase_configured(), "error": str(exc), "summary": {}, "archive_batches": []}), 400


@app.get("/api/review-queue")
def api_review_queue():
    try:
        if not supabase_configured():
            return jsonify({"enabled": False, "records": []})
        return jsonify({"enabled": True, "records": load_review_queue_from_supabase()})
    except Exception as exc:
        return jsonify({"enabled": supabase_configured(), "error": str(exc), "records": []}), 400


@app.post("/api/report")
def api_report():
    try:
        kind = request.form.get("kind", "b")
        report_start_date = report_end_date = None
        if request.form.get("report_month", ""):
            year, month = (int(part) for part in request.form.get("report_month", "").split("-"))
            report_start_date = date(year, month, 1)
            report_end_date = date(year, month, monthrange(year, month)[1])
        elif kind == "c" and (request.form.get("report_start", "") or request.form.get("report_end", "")):
            report_start_date = date.fromisoformat(request.form.get("report_start") or request.form.get("report_end"))
            report_end_date = date.fromisoformat(request.form.get("report_end") or request.form.get("report_start"))
        stays = records_to_stays(json.loads(request.form.get("stays", "[]")))
        stays = filter_stays_for_report(
            stays,
            kind,
            request.form.get("report_month", ""),
            request.form.get("report_start", ""),
            request.form.get("report_end", ""),
        )
        fee_rate = dec(request.form.get("fee_rate", "5.00"))
        settings = {k: clean(request.form.get(k, "")) for k in request.form.keys()}
        pdf = form_b_pdf(stays, settings, fee_rate, report_start_date, report_end_date) if kind == "b" else form_c_pdf(stays, settings, fee_rate, report_start_date, report_end_date)
        name = report_filename(kind, stays, report_start_date, report_end_date)
        return Response(pdf, mimetype="application/pdf", headers={"Content-Disposition": f'attachment; filename="{name}"'})
    except Exception as exc:
        return Response(str(exc), status=400)


if __name__ == "__main__":
    app.run(debug=True)
